from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.http import JsonResponse, Http404, HttpResponseRedirect, HttpResponse
from django.db.models import Q, Avg, Sum, F
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.generic import TemplateView, DetailView
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.core import serializers
from django.utils import timezone
from django.contrib.auth.models import User
# icontains query from a list
# https://thepythonguru.com/python-builtin-functions/reduce/
# https://stackoverflow.com/questions/4824759/django-query-using-contains-each-value-in-a-list
from functools import reduce
# https://docs.python.org/3/library/operator.html
# https://www.geeksforgeeks.org/reduce-in-python/
import operator
from itertools import chain
# for filtering exact salary
from decimal import Decimal

# datetime convertion
import pytz
import datetime
import time
from django.conf import settings
from django.db.models import Value
from django.db.models.functions import Concat
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color, Protection, colors, GradientFill
from openpyxl.utils import get_column_letter
import csv
import xlwt 
from cryptography.fernet import Fernet
import base64
import logging
import traceback
import xlrd 
from django.forms import modelformset_factory, inlineformset_factory
from application.forms import UserAccountCredentialsForm, PersonalForm, MobileNumberForm, SkillsForm, CompanyForm, TelephoneNumberForm, CutOffPeriodForm,AttendanceForm,AttendanceFormManual,EmployeeSalaryForm,EmployeePayrollForm,EmployeeLeavesForm, EmployeeIteneraryForm, EmployeeIteneraryDetailsForm, ConcernsEmployeeForm, ConcernsReplyEmployeeForm,OvertimeForm,OvertimeDetailsForm,RolesPermissionForm
from application.models import PersonalInfo, MobileNumberInfo, SkillsInfo, TelephoneNumberInfo, CompanyInfo, AttendanceInfo, CutOffPeriodInfo, EmployeeSalary, EmployeePayroll,EmployeeLeaves, EmployeeItenerary, EmployeeIteneraryDetails, Concerns, Notifications,Overtime,OvertimeDetails,RolesPermission
from django.forms import modelformset_factory, inlineformset_factory
from datetime import date
from django.db.models import F
# https://www.pythoncircle.com/post/641/encryption-decryption-in-python-django/
# https://www.quora.com/How-do-I-block-inspect-element-on-my-website

from application.custom_user_validator import validate

category_list = [
    'Uploading File',
    'Delete Uploaded File',
    'Creating Payroll',
    'Updating Payroll Settings',
    'Updating Profile',
    'Updating Password',
    'New User Registration',
    'Deleting Employee',
    'Reply Concerns',
    'Creating Employee Transaction', 
    'Updating Employee Transaction', 
    'Deleting Employee Transaction',
    'Add',
    'Delete',
    ]

level_list =[
        'success',
        'info',
        'warning',
        'error'
        ]

preffered_working_hours_list = [
    '8:00am-5:00pm',
    '8:30am-5:30pm',
    '9:00am-6:00pm',
]

def read_notifications(request,id):
    data = dict()
    current_user = get_object_or_404(User, username=request.user.username)
    notification = get_object_or_404(Notifications, id=id)
    if request.is_ajax():
        if request.method == 'POST':  
            if notification.url:
                data['has_url'] = True
            else: 
                data['has_url'] = False 

            notification.delete()
            
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            return JsonResponse(data)

    else:
        raise Http404()


def read_all_notifications(request):
    data = dict()
    current_user = get_object_or_404(User, username=request.user.username)
    notifications = Notifications.objects.all().filter(Q(recipient=current_user))
    if request.is_ajax():
        if request.method == 'POST': 
            for notification in notifications:
                notification.delete()
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            return JsonResponse(data)

    else:
        raise Http404()


def read_attendance_file(attendance_file, request, form, data):

    success =  True # flag for duplicate names on database 

    workbook = xlrd.open_workbook(filename=attendance_file.name, file_contents=attendance_file.read())

    schedule_information = workbook.sheet_by_index(0)
    attendance_stat = workbook.sheet_by_index(1)
    #sheetnames
    sheet_names = workbook.sheet_names()
    print(sheet_names) 
    att_log_report = workbook.sheet_by_name('Att.log report')
    # att_log_report = workbook.sheet_by_index(2)
    # For row 0 and column 0 
    cut_off_period = schedule_information.cell_value(1,1)  

    matching = CutOffPeriodInfo.objects.filter(cut_off_period=cut_off_period).count()
    if not matching > 0:
        #hold the form instance before saving for error checking
        instance = form.save(commit=False)
        instance.attendance_file = attendance_file
        instance.cut_off_period = cut_off_period
        instance.save() #save the form if successful
        

        # dates
        for index in range(3, schedule_information.ncols):
            # Days of Week
            days_of_week = schedule_information.cell_value(2,index)
            days_of_month = schedule_information.cell_value(3,index)

            # print(str(int(days_of_week)),'-', days_of_month, end=" | ")
         
            #checking if db has multiple names might duplicate the attendance record for other matching names
        starting_row = 3 # verify
        for row in range(3, att_log_report.nrows):
            
            starting_row += 2    
            starting_col = 2
            #cols
            for col in range(att_log_report.ncols):
                #rows
                
                if not starting_row > att_log_report.nrows: 
                    
                    starting_col += 1 
                    date = att_log_report.cell_value(3,col)
                    hours = att_log_report.cell_value(starting_row,col)
                    id = att_log_report.cell_value(starting_row-1,2)
                    name = att_log_report.cell_value(starting_row-1,10)   
                    # print('Row:',starting_row,'ID: ',id ,"Name: ",name)
                    timeIn = hours[:5]
                    timeOut = hours[-5:]     
                    #columns
                    if starting_col < schedule_information.ncols:   
                        days_of_month = schedule_information.cell_value(3, starting_col)    
                        if date != "": 
                            # for id
                            company_info = CompanyInfo.objects.all()   
                            for company in company_info:   
                                if not company.biometrics_id == "0":
                                    if company.biometrics_id == id:   
                                        personal = PersonalInfo.objects.filter(fk_user=company.fk_company_user).count()
                                        if personal > 0: 
                                            company_user_profile = get_object_or_404(PersonalInfo, fk_user=company.fk_company_user) 
                                            t_diff, under_time, overtime_hours = calculate_hours(timeIn, timeOut, company)
                                            # print(company.biometrics_id,'-',company.fk_company_user,'=',company_user_profile)
                                            attendance = AttendanceInfo(employee_profile=company_user_profile, cut_off_period=instance, days_of_week=days_of_month, date=str(int(date)), time_in=timeIn, time_out=timeOut, late=t_diff, undertime=under_time, overtime=overtime_hours)
                                            attendance.save() 
                            # 
                            # employees = PersonalInfo.objects.all()
                            # for employee in employees:
                            #     employee_name = '{fname} {mname} {lname}'.format(fname=employee.first_name, mname=employee.middle_name, lname=employee.last_name) 
                            #     #print(name,'-',employee_name)
                            #     if name in employee_name:
                            #         #important notes get will not return when multiple names has been detected on the database or xlsx
                            #         #employees = PersonalInfo.objects.annotate(name=Concat('first_name', Value(' '),'middle_name', Value(' '),'last_name'),).filter(Q(name__icontains=name.casefold()))
                            #         try:
                            #             employees = PersonalInfo.objects.annotate(name=Concat('first_name', Value(' '),'middle_name', Value(' '),'last_name'),).get(Q(name__icontains=name.casefold()))
                            #             attendance = AttendanceInfo(employee_profile=employees, cut_off_period=instance, days_of_week=days_of_month, date=str(int(date)), time_in=timeIn, time_out=timeOut, late=t_diff, undertime=under_time, overtime=overtime_hours)
                            #             attendance.save()
                                
                            #         except PersonalInfo.MultipleObjectsReturned:
                            #             print('Multiple Names Detected! Cannot be processed!')
                                        
                            #             success = False
                            #     else:
                            #         #print('Not Found')
                            #         pass
                            # 
                            # attendance = AttendanceInfo(employee_profile=employees, cut_off_period=instance, days_of_week=days_of_month, date=str(int(date)), time_in=timeIn, time_out=timeOut, late=t_diff, undertime=under_time)
                            # attendance.save()

                            #print(row,hours,end=" ")
                            #output = '{row} => ID: [{id}] Day:[{days}] Name: [{name}]- Date: ({date}) - Hours: ({stime}-{etime}) Late: ({dtime}) Under Time: ({utime})'.format(row=row, id=id, days=days_of_month,name=name,date=str(int(date)),stime=timeIn,etime=timeOut, dtime=t_diff, utime=under_time)
                            # output = 'ID: [{id}] Day:[{days}] Name: [{name}]- Date: ({date}) - Hours: ({stime}-{etime} O.T: ({overtime}))'.format(row=row, id=id, days=days_of_month,name=name,date=str(int(date)),stime=timeIn,etime=timeOut, overtime=overtime_hours)
                            # print(output, end=" ")
                            # print()
                            #print(row,'=>','[]',str(int(days)), '-(', hours[:5], '-', hours[-5:],')', end=" ")         
        if success:
            data['success'] = True
        else:
            data['success'] = False
            messages.warning(request, 'Multiple Names Detected! Some employee with duplicate names will not be able to process their attendance!')
        '''
        String -> json.loads() -> json Object
        json Object -> json.dumps() -> String
        '''
        url = reverse_lazy('application:delete_attendance', kwargs={'id':instance.id})
        date_time_str = str(instance.date_created)
        date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
        converted_date = date_time_obj.strftime('%b %d, %Y')
        new_record = {
            "id": instance.id,
            "cut_off": cut_off_period,
            "date_created": str(converted_date),
            "attendance_file_path": instance.attendance_file.url,
            "url": str(url),
        }
        #'{"id": "Bob", "languages": ["English", "Fench"]}'
        dict_to_string = json.dumps(new_record)
        data['list'] = json.loads(dict_to_string)
        data['form_is_valid'] = True
        # Applying notifications
        current_user = get_object_or_404(User, username=request.user.username)
        profiles = PersonalInfo.objects.all().distinct().order_by('-id')
        for p in profiles:
            Notifications.objects.create(sender=current_user,recipient=p.fk_user,message="Cut-off attendance was uploaded!",category=category_list[0],level=level_list[1])
             
    else:
        
        data['success'] = True
        data['form_is_valid'] = False
        messages.error(request, 'Record Already Exists')
   

def calculate_hours(timeIn, timeOut, company):
    t_diff = None
    under_time = None
    overtime_hours = 0
 
    working_hours_in = company.preffered_working_hours[:6]
    working_hours_out = company.preffered_working_hours[7:13]

    if timeIn and timeOut:
        # https://strftime.org/
        # watch out for time formats
        # 24 hours format from biometrics
        d_time_in = datetime.datetime.strptime(timeIn, '%H:%M')
        d_time_out = datetime.datetime.strptime(timeOut, '%H:%M')


        # 12 hours format from preferred time to work convert in the 24 hrs correct format
        d_working_hours_in = datetime.datetime.strptime(working_hours_in, '%I:%M%p')
        d_working_hours_out = datetime.datetime.strptime(working_hours_out, '%I:%M%p')

        hour = 1
        hours_added = datetime.timedelta(hours = hour) 
        #min_overtime = datetime.datetime.strptime(d_working_hours_out, '%H:%M') 
        min_overtime = d_working_hours_out + hours_added  
        if d_time_in > d_working_hours_in:
            # convert time diff to minutes 1 min = 60 secs
            t_diff =  int((d_time_in - d_working_hours_in).total_seconds() / 60.0)
            if t_diff < 0:
                t_diff = 0
        if d_time_out < d_working_hours_out:
            under_time = int((d_time_out - d_working_hours_out).total_seconds() / 60.0)
            if under_time < 0:
                under_time = 0
        
        #for ot computation
        if d_time_out > min_overtime:
            overtime_minutes = int((d_time_out - d_working_hours_out).total_seconds() / 60.0)
            overtime_hours = round(float(overtime_minutes/60),2)
            # print('[Name: ',company.fk_company_user,'] [Out: ',d_time_out,'] [W.O',d_working_hours_out,'] [O.T:',  min_overtime,'] [O.T Min: ', overtime_minutes,'] [O.T Hours: [',overtime_hours,']')
         
    
    return t_diff, under_time, overtime_hours
    

def encrypt_key(txt):
    try:
        # convert integer etc to string first
        txt = str(txt)
        # get key from settings
        cipher_suite = Fernet(settings.ENCRYPT_KEY)  # key should be byte
        # #input should be byte, so convert the text to byte
        encrypted_text = cipher_suite.encrypt(txt.encode('ascii'))
        # encode to urlsafe base64 format
        encrypted_text = base64.urlsafe_b64encode(
            encrypted_text).decode("ascii")

        return encrypted_text

    except Exception as e:
        # log if an error
        logging.getLogger("error_logger").error(traceback.format_exc())
        return None


def decrypt_key(txt):
    try:
        # base64 decode
        txt = base64.urlsafe_b64decode(txt)
        cipher_suite = Fernet(settings.ENCRYPT_KEY)
        decoded_text = cipher_suite.decrypt(txt).decode("ascii")
        return decoded_text
    except Exception as e:
        # log the error
        logging.getLogger("error_logger").error(traceback.format_exc())
        return None

# Create your views here.


def error_404(request, exception):
    template_name = "error/error_404.html"
    context = {

    }
    return render(request, template_name, context)


def error_500(request):
    template_name = "error/error_500.html"
    context = {

    }
    return render(request, template_name, context)


# class PayrollAttendanceIndexPage(LoginRequiredMixin, TemplateView):
#     template_name = "index.html"

#     def get_context_data(self, *args, **kwargs): 
#         context = super(PayrollAttendanceIndexPage, self).get_context_data(*args, **kwargs)
#         user = get_object_or_404(User, username=self.request.user.username) 


#         notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
#         notifications_count = notifications.count()
#         total_emp = PersonalInfo.objects.all().distinct().count()
#         #unapproved user registration
#         total_unpermitted_emp = User.objects.all().filter(Q(is_active=False) | Q(is_staff=False)).distinct().count()            
#         total_active_user_emp = User.objects.all().filter(Q(is_active=True) & Q(is_staff=True)).distinct().count()
#         total_attendance = AttendanceInfo.objects.all().distinct().count()  
#         total_payroll = EmployeePayroll.objects.all().distinct().count()  
#         total_leaves = EmployeeLeaves.objects.all().distinct().count()  
#         total_unapproved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Disapproved')).distinct().count()  
#         total_approved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Approved')).distinct().count()  
#         total_iteneraries = EmployeeItenerary.objects.all().distinct().count()  
#         total_unapproved_iteneraries = EmployeeItenerary.objects.all().filter(~Q(checked_by=None) & ~Q(approved_by=None)).distinct().count()  
#         total_approved_iteneraries = EmployeeItenerary.objects.all().filter(Q(checked_by=None) & Q(approved_by=None)).distinct().count() 
#         total_concern =  Concerns.objects.all().distinct().count() 
#         total_unreplied_concern = Concerns.objects.all().filter(Q(reply=None)).distinct().count()   
#         emp_payrolls = EmployeePayroll.objects.all()
#         year_list = [] 

#         yearly_employee_wages_list = []
#         yearly_emplpyee_deduction_list = []

#         for emp in emp_payrolls:
#             year = str(emp.payroll_cutoff_period)[:4]
#             if not year in year_list:
#                 year_list.append(year)  
        
#         #print(year_list)

#         for year in year_list:
#             total_net_pay = 0
#             total_deductions = 0
#             cutoffperiod = CutOffPeriodInfo.objects.all().filter(Q(cut_off_period__icontains=year)) 
#             for cutoff in cutoffperiod:
#                 emp_payroll = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('net_pay')).get('net_pay__sum') 
#                 emp_deductions  = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('total_deduction')).get('total_deduction__sum') 
#                 if emp_payroll:
#                     total_net_pay = total_net_pay + emp_payroll 
#                     total_deductions = total_deductions + emp_deductions
#                 #print(year,'=',cutoff,':',emp_payroll)
#             #print('----')
#             #print(year,'=',total_net_pay)
            
#             yearly_employee_wages_list.append({'year': year, 'value': int(round(total_net_pay,2))})
#             yearly_emplpyee_deduction_list.append({'year': year, 'deduction': int(round(total_deductions,2)), 'wages': int(round(total_net_pay,2))})

#         dumps_yearly_employee_wages_list = json.dumps(yearly_employee_wages_list)
#         loads_yearly_employee_wages_list =json.loads(dumps_yearly_employee_wages_list)
#         # print(dumps_yearly_employee_wages_list)
        
#         dumps_yearly_employee_deduction_list = json.dumps(yearly_emplpyee_deduction_list)
#         loads_yearly_employee_deduction_list =json.loads(dumps_yearly_employee_deduction_list)
                
#         # https://www.educative.io/edpresso/what-is-the-difference-between-jsonloads-and-jsondumps

#         context.update({
#             'user': user, 
#             'total_emp': total_emp,
#             'total_unpermitted_emp': total_unpermitted_emp,
#             'total_active_user_emp': total_active_user_emp,
#             'total_attendance': total_attendance,
#             'total_payroll': total_payroll,
#             'total_leaves': total_leaves,
#             'total_unapproved_leaves': total_unapproved_leaves,
#             'total_approved_leaves': total_approved_leaves,
#             'total_iteneraries': total_iteneraries,
#             'total_unapproved_iteneraries': total_unapproved_iteneraries,
#             'total_approved_iteneraries': total_approved_iteneraries,
#             'total_concern': total_concern,
#             'total_unreplied_concern': total_unreplied_concern, 
#             'loads_yearly_employee_wages_list': loads_yearly_employee_wages_list,
#             'loads_yearly_employee_deduction_list': loads_yearly_employee_deduction_list,
#             'notifications': notifications,
#             'notifications_count': notifications_count,
#         })
#         return context
         
            

@login_required
def PayrollAttendanceIndexPage(request):
    template_name = "index.html"
    user = get_object_or_404(User, username=request.user.username)  
    user = validate(user,is_managing_director=True, is_human_resource=True)

    if user:  
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        total_emp = PersonalInfo.objects.all().distinct().count()
        #unapproved user registration
        total_unpermitted_emp = User.objects.all().filter(Q(is_active=False) | Q(is_staff=False)).distinct().count()            
        total_active_user_emp = User.objects.all().filter(Q(is_active=True) & Q(is_staff=True)).distinct().count()
        total_attendance = AttendanceInfo.objects.all().distinct().count()  
        total_payroll = EmployeePayroll.objects.all().distinct().count()  
        total_leaves = EmployeeLeaves.objects.all().distinct().count()  
        total_unapproved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Disapproved')).distinct().count()  
        total_approved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Approved')).distinct().count()  
        total_iteneraries = EmployeeItenerary.objects.all().distinct().count()  
        total_unapproved_iteneraries = EmployeeItenerary.objects.all().filter(~Q(checked_by=None) & ~Q(approved_by=None)).distinct().count()  
        total_approved_iteneraries = EmployeeItenerary.objects.all().filter(Q(checked_by=None) & Q(approved_by=None)).distinct().count() 
        total_concern =  Concerns.objects.all().distinct().count() 
        total_unreplied_concern = Concerns.objects.all().filter(Q(reply=None)).distinct().count()   
        emp_payrolls = EmployeePayroll.objects.all()
        year_list = [] 

        yearly_employee_wages_list = []
        yearly_emplpyee_deduction_list = []

        for emp in emp_payrolls:
            year = str(emp.payroll_cutoff_period)[:4]
            if not year in year_list:
                year_list.append(year)  
        
        #print(year_list)

        for year in year_list:
            total_net_pay = 0
            total_deductions = 0
            cutoffperiod = CutOffPeriodInfo.objects.all().filter(Q(cut_off_period__icontains=year)) 
            for cutoff in cutoffperiod:
                emp_payroll = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('net_pay')).get('net_pay__sum') 
                emp_deductions  = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('total_deduction')).get('total_deduction__sum') 
                if emp_payroll:
                    total_net_pay = total_net_pay + emp_payroll 
                    total_deductions = total_deductions + emp_deductions
                #print(year,'=',cutoff,':',emp_payroll)
            #print('----')
            #print(year,'=',total_net_pay)
            
            yearly_employee_wages_list.append({'year': year, 'value': int(round(total_net_pay,2))})
            yearly_emplpyee_deduction_list.append({'year': year, 'deduction': int(round(total_deductions,2)), 'wages': int(round(total_net_pay,2))})

        dumps_yearly_employee_wages_list = json.dumps(yearly_employee_wages_list)
        loads_yearly_employee_wages_list =json.loads(dumps_yearly_employee_wages_list)
        # print(dumps_yearly_employee_wages_list)
        
        dumps_yearly_employee_deduction_list = json.dumps(yearly_emplpyee_deduction_list)
        loads_yearly_employee_deduction_list =json.loads(dumps_yearly_employee_deduction_list)
                
        # https://www.educative.io/edpresso/what-is-the-difference-between-jsonloads-and-jsondumps

        context = {
            'user': user, 
            'total_emp': total_emp,
            'total_unpermitted_emp': total_unpermitted_emp,
            'total_active_user_emp': total_active_user_emp,
            'total_attendance': total_attendance,
            'total_payroll': total_payroll,
            'total_leaves': total_leaves,
            'total_unapproved_leaves': total_unapproved_leaves,
            'total_approved_leaves': total_approved_leaves,
            'total_iteneraries': total_iteneraries,
            'total_unapproved_iteneraries': total_unapproved_iteneraries,
            'total_approved_iteneraries': total_approved_iteneraries,
            'total_concern': total_concern,
            'total_unreplied_concern': total_unreplied_concern, 
            'loads_yearly_employee_wages_list': loads_yearly_employee_wages_list,
            'loads_yearly_employee_deduction_list': loads_yearly_employee_deduction_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

 

def login_page(request):
    template_name = "registration/login.html"

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)

        if user:
            if user.is_active and user.is_staff and user.is_superuser:
                # request.session.set_expiry(300)
                login(request, user)
                logged_in_time = get_object_or_404(User, username=user)
                cd = datetime.datetime.now()
                print('==========================', cd)
                cd = timezone.now()
                print('--------------------------', cd)
                logged_in_time.last_login = cd
                logged_in_time.save()
                return HttpResponseRedirect(reverse("index"))
            elif user.is_active and user.is_staff and not user.is_superuser: 
                # request.session.set_expiry(300)
                login(request, user)
                logged_in_time = get_object_or_404(User, username=user)
                cd = datetime.datetime.now()
                print('==========================', cd)
                cd = timezone.now()
                print('--------------------------', cd)
                logged_in_time.last_login = cd
                logged_in_time.save()
                return HttpResponseRedirect(reverse("employee_index"))

            else:
                messages.warning(
                    request, "Your account is not active! please contact your administrator.")
        else:
            messages.error(
                request, "Your account is INVALID! please try again.")
    elif request.method == 'GET':
        storage = messages.get_messages(request)
        storage.used = True

    return render(request, template_name)


def register_page(request):
    template_name = "registration/register.html"
    if request.method == 'POST':
        form = UserAccountCredentialsForm(request.POST or None)
        if form.is_valid():
            userForm = form.save(commit=False)
            userForm.is_active = False
            userForm.is_staff = False
            userForm.is_superuser = False
            userForm.save()
            messages.success(
                request, "You have successfully registered, please try to login.")
            return HttpResponseRedirect(reverse("login"))
        else:
            messages.error(request, "Error")
    elif request.method == 'GET':
        form = UserAccountCredentialsForm(request.GET or None)
    context = {
        'form': form,
    }
    return render(request, template_name, context)


@login_required
def user_profile(request):
    # issue
    template_name = "profiles/user_profile.html"
    user = get_object_or_404(User, username=request.user.username)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    #notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    print('-------------->',user)
    # When using in template it was automatically formatted
    # https://ourcodeworld.com/articles/read/555/how-to-format-datetime-objects-in-the-view-and-template-in-django
    last_time_logged = user.last_login
    last_time_joined = user.date_joined
    # last_time_logged = str(user.last_login)[:-6]
    # last_time_joined = str(user.date_joined)[:-6]
    # converted_last_time_logged = datetime.strptime(last_time_logged, '%Y-%m-%d %H:%M:%S.%f')
    # formatted_last_time_logged = datetime.strftime(converted_last_time_logged, '%m/%d/%Y %I:%M:%S %p')

    # converted_last_time_joined = datetime.strptime(last_time_joined, '%Y-%m-%d %H:%M:%S.%f')
    # formatted_last_time_joined = datetime.strftime(converted_last_time_joined, '%m/%d/%Y %I:%M:%S %p')
    # datetime_object = datetime.strptime(str(user.last_login), "%Y-%m-%dT%H:%M:%S.%f%Z")
    # print('-------------------------',datetime_object)
    if user.is_active and user.is_staff and user.is_superuser:
        personalInfoFormset = inlineformset_factory(
            User, PersonalInfo, form=PersonalForm, extra=0, can_delete=False)
        if request.method == "GET":
            piformset = personalInfoFormset(request.GET or None, instance=user)
            record = PersonalInfo.objects.filter(fk_user=user)
            company = CompanyInfo.objects.filter(fk_company_user=user)
            # profile = get_object_or_404(PersonalInfo, fk_user=user)
            # company_details = get_object_or_404(CompanyInfo, fk_company_user=user)
            profiles =  PersonalInfo.objects.all().filter(fk_user=user).distinct()
            company_details = CompanyInfo.objects.all().filter(fk_company_user=user).distinct()
            print(profiles)
        elif request.method == "POST":
            pass

        context = {
            'user': user,
            'last_time_logged': last_time_logged,
            'last_time_joined': last_time_joined,
            'piformset': piformset,
            'record': record,
            'profiles': profiles,
            'company':company,
            'company_details':company_details,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def new_profile(request):
    data = dict()
    template_name = "profiles/new_profile_personal.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = PersonalForm(request.GET or None)
        elif request.method == 'POST':
            form = PersonalForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                print('---------------------------------------------->',instance.id)
                instance.fk_user = current_user
                #issue is cannot retrieve id to encrypt before saving 
                #instance.key_id = encrypt_key(instance.id)
                instance.save() 
                instance.key_id = encrypt_key(instance.id)
                instance.save() 
                print('==============================================>',instance.id)
                Notifications.objects.create(sender=current_user,recipient=current_user,message="You have been created your new profile!",category=category_list[4],level=level_list[1])
                
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

# On progress
@login_required
def edit_profile(request):
    data = dict()
    template_name = "profiles/edit_profile_personal.html"
    current_user = get_object_or_404(User, username=request.user.username)
    profile = get_object_or_404(PersonalInfo, fk_user=current_user)
    if request.is_ajax():
        if request.method == 'GET':

            form = PersonalForm(request.GET or None, instance=profile)

        elif request.method == 'POST':
            form = PersonalForm(request.POST or None, instance=profile)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.key_id = encrypt_key(profile.id)
                instance.save() 
                Notifications.objects.create(sender=current_user,recipient=current_user,message="You have been updated your profile!",category=category_list[4],level=level_list[1])
                
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


def save_profile_image(request):
    data = dict()
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'POST':
            image = request.FILES['image']
            personal_info, is_created = PersonalInfo.objects.update_or_create(
                fk_user=current_user, defaults={'image': image, })
            if is_created:
                print("Already Exists!")
            else:
                print("New Record!")
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Your profile picture has been updated!",category=category_list[4],level=level_list[1])
            
            data['is_save'] = True
        return JsonResponse(data)
    else:
        raise Http404()


@login_required
def add_mobile_number(request):
    data = dict()
    template_name = "profiles/add_mobile_number.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = MobileNumberForm(request.GET or None)
        elif request.method == 'POST':
            form = MobileNumberForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_mobile_user = current_user
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="New mobile number has been added!",category=category_list[12],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False
        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


def delete_mobile_number(request, id):
    data = dict()
    mobile_id = get_object_or_404(MobileNumberInfo, id=id)
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'POST':
            print(mobile_id.mobile_number)

            mobile_id.delete()
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Mobile number has been deleted!",category=category_list[13],level=level_list[1])
            
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            return JsonResponse(data)
    else:
        raise Http404()

def add_skill(request):
    data = dict()
    template_name = "profiles/add_skills.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = SkillsForm(request.GET or None)
        elif request.method == 'POST':
            form = SkillsForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_skills_user = current_user
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="New Skill has been added!",category=category_list[12],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False
        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

def delete_skill(request, id):
    data = dict()
    skill_id = get_object_or_404(SkillsInfo, id=id)
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'POST':
            skill_id.delete()
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Skill has been Deleted!",category=category_list[13],level=level_list[1])
            
            return JsonResponse(data)
    else:
        raise Http404()

    return JsonResponse(data)

    # return HttpResponseRedirect(reverse('application:user_profile'))

@login_required
def new_company_details(request):
    data = dict()
    template_name = "profiles/new_company_details.html"
    current_user = get_object_or_404(User, username=request.user.username)

    if request.is_ajax():
        if request.method == 'GET':
            form = CompanyForm(request.GET or None)
        elif request.method == 'POST':
            form = CompanyForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_company_user = current_user 
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="Company profile has been added!",category=category_list[4],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                print(form.errors)
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(
            template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def edit_company_details(request):
    data = dict()
    template_name = "profiles/edit_company_details.html"
    current_user = get_object_or_404(User, username=request.user.username)
    company = get_object_or_404(CompanyInfo, fk_company_user=current_user)
    if request.is_ajax():
        if request.method == 'GET':

            form = CompanyForm(request.GET or None, instance=company)

        elif request.method == 'POST':
            form = CompanyForm(request.POST or None, instance=company)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="Company profile has been updated!",category=category_list[4],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def add_telephone_number(request):
    data = dict()
    template_name = "profiles/add_telephone_number.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = TelephoneNumberForm(request.GET or None)
        elif request.method == 'POST':
            form = TelephoneNumberForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_telephone_user = current_user
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="New Telephone number has been added!",category=category_list[12],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False
        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

def delete_telephone_number(request, id):
    data = dict()
    telephone_id = get_object_or_404(TelephoneNumberInfo, id=id)

    if request.is_ajax():
        if request.method == 'POST': 
            telephone_id.delete()
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Telephone number has been Deleted!",category=category_list[13],level=level_list[1])
            
            return JsonResponse(data)
    else:
        raise Http404()

@login_required 
def cut_off_page(request):
    template_name = "attendance/cut_off_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

        context = {
            'user':user, 
            'cut_off_period_list': cut_off_period_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

def delete_attendance(request, id):
    data = dict()
    template_name = 'attendance/delete_attendance.html'
    record = get_object_or_404(CutOffPeriodInfo, id=id)
    if request.is_ajax():
        if request.method == 'GET':
            context = {
                'record': record,
            }
            data['html_form'] = render_to_string(template_name, context, request)
        elif request.method == 'POST':
            cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
            record.delete()
            path = record.attendance_file.path
            '''
            path. exists(path) - Returns true if the path is a file, directory, or a valid symlink.
            path. isfile(path) - Returns true if the path is a regular file or a symlink to a file.
            path. isdir(path) - Returns true if the path is a directory or a symlink to a directory
            ''' 
            if os.path.isfile(path):
                os.remove(path)
                print('file path is existing')
            else:
                print('file does not exists')

            context = {
                'cut_off_period_list': cut_off_period_list,
            }
            
            data['form_is_valid'] = True
            data['table_records'] = render_to_string('attendance/cut_off_page_table.html', context) 
            # Applying notifications
            current_user = get_object_or_404(User, username=request.user.username)
            profiles = PersonalInfo.objects.all().distinct().order_by('-id')
            for p in profiles:
                Notifications.objects.create(sender=current_user,recipient=p.fk_user,message="Cut-off attendance was deleted!",category=category_list[1],level=level_list[1])
            
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_list_page(request):
    template_name = "employees/employee_list.html"
    user = get_object_or_404(User, username=request.user.username)
    
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        employees_company = User.objects.all().filter(Q(is_active=True) & Q(is_staff=True) ).distinct()
        employees = PersonalInfo.objects.all().filter(fk_user__in=employees_company).exclude(Q(key_id='') | Q(first_name='') | Q(middle_name='') | Q(last_name='last_name'))
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        context = {
            'user':user, 
            'employees':employees,
            'employees_company': employees_company,
            'notifications': notifications,
            'notifications_count': notifications_count,
        } 
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))     

@login_required
def employee_view_profile(request, key):
    template_name = "employees/employee_view_employee_profiles.html"
    user = get_object_or_404(User, username=request.user.username)
    id = decrypt_key(key) 

    personal_info = get_object_or_404(PersonalInfo, id=id)
    target_user_parent_user = get_object_or_404(User, profile_to_user=personal_info)#look up from child to parent model
    
    company_info = CompanyInfo.objects.filter(fk_company_user=target_user_parent_user)
    mobile_info = MobileNumberInfo.objects.filter(fk_mobile_user=target_user_parent_user)
    telephone_info = TelephoneNumberInfo.objects.filter(fk_telephone_user=target_user_parent_user)
    skills_info = SkillsInfo.objects.filter(fk_skills_user=target_user_parent_user)

    last_time_logged = target_user_parent_user.last_login
    last_time_joined = target_user_parent_user.date_joined

    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and user.is_superuser:
        context = {
            'user': user,
            'preffered_working_hours_list':preffered_working_hours_list,
            'last_time_logged': last_time_logged,
            'last_time_joined': last_time_joined,
            'personal_info':personal_info, 
            'company_info': company_info,
            'mobile_info': mobile_info,
            'telephone_info':telephone_info,
            'skills_info':skills_info,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def set_employee_preferred_working_hours(request, id):
    data = dict()
    user = get_object_or_404(User, username=request.user.username)
    company = get_object_or_404(CompanyInfo, id=id)
    if request.is_ajax():
        # https://www.educative.io/edpresso/what-is-the-difference-between-jsonloads-and-jsondumps
        '''
        json objects are surrounded by curly braces { }. They are written in key and value pairs.

        json.loads() takes in a string and returns a json object.

        json.dumps() takes in a json object and returns a string.

        As you can see, json.dumps() and json.loads() are opposite of one another.
        '''
        if request.method == 'GET':
            pass 
        elif request.method == 'POST':
            data_from_js = json.loads(request.body)
            print(json.dumps(data_from_js, indent=4, sort_keys=True))

            value = data_from_js['value']
            print(value,company)

            new, existing = CompanyInfo.objects.update_or_create(id=id, defaults={
                'preffered_working_hours': value,
            })

        data['context'] = data_from_js

        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def delete_employee(request, id):
    data = dict()
    template_name = "employees/delete_employee.html"
    personal_info = get_object_or_404(PersonalInfo, id=id)
    current_user = get_object_or_404(User, profile_to_user=personal_info)
    if request.is_ajax():
        if request.method == 'GET':
            
            context = {
            'personal_info': personal_info,
            'user':current_user,
            }
            data['html_form'] = render_to_string(template_name, context, request)

        elif request.method == 'POST':
            current_user.delete()
            data['form_is_valid'] = True 
            profiles = PersonalInfo.objects.all().distinct().order_by('-id')
            for p in profiles:
                Notifications.objects.create(sender=current_user,recipient=p.fk_user,message="An employee was deleted!",category=category_list[7],level=level_list[1])
            
  
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_attendance(request, pk):
    template_name = "employees/employee_attendance_list.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        id = decrypt_key(pk)
        employee = get_object_or_404(PersonalInfo, id=id)
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
        # attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period__in=cut_off_period_list)).distinct()
        # print(attendance)
        context = {
            'employee':employee,
            'user':user, 
            'cut_off_period_list': cut_off_period_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
       
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_attendance_cutoff(request, pk, id):
    template_name = "employees/employee_cutoff_page.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        pid = decrypt_key(pk)
        employee = get_object_or_404(PersonalInfo, id=pid) 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

        today = date.today()
        year_today = today.strftime("%Y") 

        attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period=cutoff)).distinct()
        #print(attendance,cutoff)

        AttendanceInlineFormset = inlineformset_factory(PersonalInfo, AttendanceInfo, form=AttendanceForm, extra=0, can_delete=False)
        
        is_available = True
        selected_cut_off = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
        
        
        if not selected_cut_off:
            is_available = False             
        
        if request.method == 'GET':
            formset = AttendanceInlineFormset(request.GET or None, instance=employee, queryset=AttendanceInfo.objects.filter(cut_off_period=cutoff).order_by("id"))
            
        elif request.method == 'POST':
            formset = AttendanceInlineFormset(request.POST or None, instance=employee, queryset=AttendanceInfo.objects.filter(cut_off_period=cutoff).order_by("id"))
            if formset.is_valid(): 
                #https://docs.djangoproject.com/es/2.1/topics/forms/#looping-over-hidden-and-visible-fields
                """
                If you’re manually laying out a form in a template, as opposed to relying on Django’s default form layout, 
                you might want to treat <input type="hidden"> fields differently from non-hidden fields. 
                For example, because hidden fields don’t display anything, putting error messages «next to» the field 
                could cause confusion for your users – so errors for those fields should be handled differently.

                Django provides two methods on a form that allow you to loop over the hidden and visible fields 
                independently: hidden_fields() and visible_fields(). Here’s a modification of an earlier example that uses 
                these two methods:

                {# Include the hidden fields #}
                {% if form in formset.forms %}
                    {% for hidden in form.hidden_fields %}
                        {{ hidden }}    
                    {% endfor %}
                {% endif %}            
                {# Include the visible fields #}
                {% for field in form.visible_fields %}
                    <div class="fieldWrapper">
                        {{ field.errors }}
                        {{ field.label_tag }} {{ field }}
                    </div>
                {% endfor %}

                This example does not handle any errors in the hidden fields. Usually, an error in a hidden field is a 
                sign of form tampering, since normal form interaction won’t alter them. However, you could easily insert 
                some error displays for those form errors, as well.

                """
                # Hidden id causes error id is required
                # for form in formset.forms: 
                #     for hidden in form.hidden_fields():                    
                #         print(hidden)
                
                # form = formset.save(commit=False)
                # form.save()
                myform = formset.save(commit=False)
                for mf in myform:
                    mf.employee_profile = employee
                    mf.save()
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_cutoff_page', kwargs={'id':cutoff.id}))
                Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your attendance for cut-off {cutoffperiod} was created!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])
                return HttpResponseRedirect(reverse_lazy('application:employee_attendance_view', kwargs={'pk':employee.key_id}))
            else:
                messages.error(request, "Form Error")
        context = {
            'formset':formset,
            'is_available':is_available,
            'employee':employee,
            'user':user,  
            'cutoff': cutoff,
            'notifications': notifications,
            'notifications_count': notifications_count,
            'year_today': year_today,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
		
@login_required
def employee_attendance_manual_configuration(request, pk, id):
    template_name = "employees/employee_cutoff_manual_page.html"
    user = get_object_or_404(User, username=request.user.username)     
    # working_hours_in = company.preffered_working_hours[:6]
    # working_hours_out = company.preffered_working_hours[7:13]  
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        pid = decrypt_key(pk)
        employee = get_object_or_404(PersonalInfo, id=pid) 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        today = date.today()
        year_today = today.strftime("%Y") 
        AttendanceInlineFormset = inlineformset_factory(CutOffPeriodInfo, AttendanceInfo, form=AttendanceFormManual, extra=16, can_delete=False) 
        try:  
            # https://ourcodeworld.com/articles/read/555/how-to-format-datetime-objects-in-the-view-and-template-in-django
            # https://adamj.eu/tech/2020/02/18/safely-including-data-for-javascript-in-a-django-template/
            # https://docs.djangoproject.com/en/3.1/ref/templates/builtins/ 
            company = CompanyInfo.objects.get(fk_company_user=employee.fk_user) 
            # 12 hrs format by defaut
            working_hours_in = company.preffered_working_hours[:6]
            working_hours_out = company.preffered_working_hours[7:13]
            # parsing of time from 12 hrs to 24 hrs format I = (12hrs format)
            working_hours_in = datetime.datetime.strptime(working_hours_in, '%I:%M%p')
            working_hours_out = datetime.datetime.strptime(working_hours_out, '%I:%M%p')
            # formating of time H = (24hrs format)
            working_hours_in = datetime.datetime.strftime(working_hours_in, '%H:%M')
            working_hours_out = datetime.datetime.strftime(working_hours_out, '%H:%M')

            preferred_working_hours_data = {
                'preferred_time_in': working_hours_in,
                'preferred_time_out': working_hours_out,
            }

            if request.method == 'GET':
                formset = AttendanceInlineFormset(request.GET or None, instance=cutoff, queryset=AttendanceInfo.objects.filter(employee_profile=employee).order_by("id"))        
            elif request.method == 'POST':
                formset = AttendanceInlineFormset(request.POST or None, instance=cutoff, queryset=AttendanceInfo.objects.filter(employee_profile=employee).order_by("id"))
                if formset.is_valid():
                    myform = formset.save(commit=False)
                    for mf in myform:
                        mf.employee_profile = employee
                        mf.cut_off_period = cutoff
                        mf.save()
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_cutoff_page', kwargs={'id':cutoff.id}))
                    Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your attendance for cut-off {cutoffperiod} was manually created!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])
                    return HttpResponseRedirect(reverse_lazy('application:employee_attendance_view', kwargs={'pk':employee.key_id}))
                
                else:
                    messages.error(request, "Form Error")
            
            context = {
            'formset':formset,
            'employee':employee,
            'user':user,  
            'cutoff': cutoff,
            'notifications': notifications,
            'notifications_count': notifications_count,
            'year_today':year_today,
            'preferred_working_hours_data':preferred_working_hours_data,
            }
            return render(request, template_name, context) 
        except CompanyInfo.DoesNotExist:
            return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
				 
@login_required
def employee_payroll_page(request):
    template_name = "payroll/employee_manage_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)

     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        context = {
        'user':user, 
        'cut_off_period_list':cut_off_period_list,
        'notifications': notifications,
        'notifications_count': notifications_count,
        } 
        return render(request, template_name, context)

    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  
    
@login_required
def employee_payroll_summary_page(request, id):
    template_name = "payroll/employee_payroll_summary.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        cut_off_period_list = EmployeePayroll.objects.all().filter(payroll_cutoff_period=cutoff).order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

        payroll_list = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).order_by('-id').distinct()

        t_sms = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('basic_pay')).get('basic_pay__sum')
        t_oth = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('ot_hours')).get('ot_hours__sum')
        t_ota = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('ot_pay')).get('ot_pay__sum')
        t_hol = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('holiday_pay')).get('holiday_pay__sum')
        t_ald = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('absences')).get('absences__sum')
        t_alm = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(total=Sum(F('late_min') + F('undertime_min')))['total']
        t_ala = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(total=Sum(F('late_undertime_min_amount') + F('absences_amount')))['total']
        t_gs = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('gross_pay')).get('gross_pay__sum')
        t_dp = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('philhealth_contribution')).get('philhealth_contribution__sum')
        t_ds = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('sss_premiums')).get('sss_premiums__sum')
        t_dh = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('pagibig_contribution')).get('pagibig_contribution__sum')
        t_dt = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('withholding_tax')).get('withholding_tax__sum')
        t_dl = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('deducted_salary_cash_advance')).get('deducted_salary_cash_advance__sum')
        t_do = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('pagibig_loan')).get('pagibig_loan__sum')
        t_td = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('total_deduction')).get('total_deduction__sum')
        t_npa = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('net_pay')).get('net_pay__sum')
        context = {
            'user':user,  
            'payroll_list':payroll_list,
            'cutoff': cutoff,
            't_sms':t_sms,
            't_oth':t_oth,
            't_ota':t_ota,
            't_hol':t_hol,
            't_ald':t_ald,
            't_alm':t_alm,
            't_ala':t_ala,
            't_gs':t_gs,
            't_dp':t_dp,
            't_ds':t_ds,
            't_dh':t_dh,
            't_dt':t_dt,
            't_dl':t_dl,
            't_do':t_do,
            't_td':t_td,
            't_npa':t_npa,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_payroll_summary_print_page(request, id):
    template_name = "payroll/employee_payroll_summary_print.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        cut_off_period_list = EmployeePayroll.objects.all().filter(payroll_cutoff_period=cutoff).order_by('-id').distinct()

        payroll_list = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).order_by('-id').distinct()

        t_sms = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('basic_pay')).get('basic_pay__sum')
        t_oth = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('ot_hours')).get('ot_hours__sum')
        t_ota = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('ot_pay')).get('ot_pay__sum')
        t_hol = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('holiday_pay')).get('holiday_pay__sum')
        t_ald = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('absences')).get('absences__sum')
        t_alm = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(total=Sum(F('late_min') + F('undertime_min')))['total']
        t_ala = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(total=Sum(F('late_undertime_min_amount') + F('absences_amount')))['total']
        t_gs = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('gross_pay')).get('gross_pay__sum')
        t_dp = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('philhealth_contribution')).get('philhealth_contribution__sum')
        t_ds = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('sss_premiums')).get('sss_premiums__sum')
        t_dh = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('pagibig_contribution')).get('pagibig_contribution__sum')
        t_dt = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('withholding_tax')).get('withholding_tax__sum')
        t_dl = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('deducted_salary_cash_advance')).get('deducted_salary_cash_advance__sum')
        t_do = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('pagibig_loan')).get('pagibig_loan__sum')
        t_td = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('total_deduction')).get('total_deduction__sum')
        t_npa = EmployeePayroll.objects.filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('net_pay')).get('net_pay__sum')

        context = {
        'user':user,  
        'payroll_list':payroll_list,
        'cutoff': cutoff,
        't_sms':t_sms,
        't_oth':t_oth,
        't_ota':t_ota,
        't_hol':t_hol,
        't_ald':t_ald,
        't_alm':t_alm,
        't_ala':t_ala,
        't_gs':t_gs,
        't_dp':t_dp,
        't_ds':t_ds,
        't_dh':t_dh,
        't_dt':t_dt,
        't_dl':t_dl,
        't_do':t_do,
        't_td':t_td,
        't_npa':t_npa, 
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
	
@login_required
def employee_payroll_employee_list(request, id):
    template_name = "payroll/employee_manage_payroll_employee_list.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        employees_company = User.objects.all().filter(Q(is_active=True) & Q(is_staff=True)).distinct()
        employees = PersonalInfo.objects.all().filter(fk_user__in=employees_company).exclude(Q(key_id='') | Q(first_name='') | Q(middle_name='') | Q(last_name='last_name'))
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

        context = {
            'user':user, 
            'cutoff': cutoff,
            'employees':employees,
            'employees_company': employees_company,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
 
 
def sss_contribution(monthy_salary_credit):
    #https://sssinquiries.com/contributions/new-sss-contributions-table-and-payment-schedule-2019/
    if monthy_salary_credit < 2250:
        return 80
    elif monthy_salary_credit >= 2250 and monthy_salary_credit <= 2749.99:
        return 100
    elif monthy_salary_credit >= 2750 and monthy_salary_credit <= 3249.99:
        return 120
    elif monthy_salary_credit >= 3250 and monthy_salary_credit <= 3749.99:
        return 140
    elif monthy_salary_credit >= 3750 and monthy_salary_credit <= 4249.99:
        return 160
    elif monthy_salary_credit >= 4250 and monthy_salary_credit <= 4749.99:
        return 180
    elif monthy_salary_credit >= 4750 and monthy_salary_credit <= 5249.99:
        return 200
    elif monthy_salary_credit >= 5250 and monthy_salary_credit <= 5749.99:
        return 220
    elif monthy_salary_credit >= 5750 and monthy_salary_credit <= 6249.99:
        return 240
    elif monthy_salary_credit >= 6250 and monthy_salary_credit <= 6749.99:
        return 260
    elif monthy_salary_credit >= 6750 and monthy_salary_credit <= 7249.99:
        return 280
    elif monthy_salary_credit >= 7250 and monthy_salary_credit <= 7749.99:
        return 300
    elif monthy_salary_credit >= 7750 and monthy_salary_credit <= 8249.99:
        return 320
    elif monthy_salary_credit >= 8250 and monthy_salary_credit <= 8749.99:
        return 340
    elif monthy_salary_credit >= 8750 and monthy_salary_credit <= 9249.99:
        return 360
    elif monthy_salary_credit >= 9250 and monthy_salary_credit <= 9749.99:
        return 380
    elif monthy_salary_credit >= 9750 and monthy_salary_credit <= 10249.99:
        return 400
    elif monthy_salary_credit >= 10250 and monthy_salary_credit <= 10749.99:
        return 420
    elif monthy_salary_credit >= 10760 and monthy_salary_credit <= 11249.99:
        return 440
    elif monthy_salary_credit >= 11250 and monthy_salary_credit <= 11749.99:
        return 460
    elif monthy_salary_credit >= 11750 and monthy_salary_credit <= 12249.99:
        return 480
    elif monthy_salary_credit >= 12250 and monthy_salary_credit <= 12749.99:
        return 500
    elif monthy_salary_credit >= 12760 and monthy_salary_credit <= 13249.99:
        return 520
    elif monthy_salary_credit >= 13250 and monthy_salary_credit <= 13749.99:
        return 540
    elif monthy_salary_credit >= 13750 and monthy_salary_credit <= 14249.99:
        return 560
    elif monthy_salary_credit >= 14250 and monthy_salary_credit <= 14749.99:
        return 580
    elif monthy_salary_credit >= 14750 and monthy_salary_credit <= 15249.99:
        return 600
    elif monthy_salary_credit >= 15250 and monthy_salary_credit <= 15749.99:
        return 620
    elif monthy_salary_credit >= 15750 and monthy_salary_credit <= 16249.99:
        return 640
    elif monthy_salary_credit >= 16250 and monthy_salary_credit <= 16749.99:
        return 660
    elif monthy_salary_credit >= 16750 and monthy_salary_credit <= 17249.99:
        return 680
    elif monthy_salary_credit >= 17250 and monthy_salary_credit <= 17749.99:
        return 700
    elif monthy_salary_credit >= 17750 and monthy_salary_credit <= 18249.99:
        return 720
    elif monthy_salary_credit >= 18250 and monthy_salary_credit <= 18749.99:
        return 740
    elif monthy_salary_credit >= 18750 and monthy_salary_credit <= 19249.99:
        return 760
    elif monthy_salary_credit >= 19250 and monthy_salary_credit <= 19749.99:
        return 780
    elif monthy_salary_credit >= 19750:
        return 800 

def phlhealth_contribution(monthy_salary_credit):
    ph_contrib = float(monthy_salary_credit) * 0.03
    return ph_contrib
   
def day_shift_payroll_computation_v1_2(data):
    WORKING_DAYS = [
        "MON","TUE","WED","THU","FRI", 
    ]
    REST_DAYS = [
       "SAT", "SUN",
    ]
    PAYROLL_INFO ={
        # "total_overtime": 0,
        # "total_absences": 0,
        "OT_HOURS": 0,
        "LATE_MIN": 0,
        "UNDERTIME_MIN": 0,
        "LATE_UNDERTIME_AMOUNT": 0,
        "ABSENCES": 0,
        "ABSENCES_AMOUNT": 0,
        "OT_PAY": 0,
        "HOLIDAY_PAY": 0,
    }
    OVERTIME_CATEGORY = [
        "No Overtime",
        "Regular Day",
        "Rest Day",
    ]
    HOLIDAY_CATEGORY =[
        "No Holiday",
        "Regular Holiday",
        "Special Holiday",
    ]
    total_overtime = 0 # total overtime
    total_late_absences = 0

    #Overtime per hour
    OT_Regular_Day = 1.25
    OT_Rest_Day = 1.30

    #Holiday Pay Per Day
    Regular_Holiday_Pay = 2.00
    Special_Holiday_Pay = 0.30

    monthly_rate = data["employee_salary"]
    monthly_allowance = data["employee_allowance"]
    #daily_rate = (monthly_rate * 12) / 313
    # daily_rate = (monthly_rate * 12) / 261 #because based on the payroll 5 days a week is counting
    # hourly_rate = daily_rate / 8
    # min_rate = hourly_rate / 60

    daily_rate = monthly_rate / 22
    hourly_rate = daily_rate / 8
    min_rate = hourly_rate / 60

    # print("DR: ",daily_rate)
    # print("HR: ", hourly_rate)
    # print("MR: ",min_rate)

    regular_hourly_rate_overtime = float(hourly_rate) * float(OT_Regular_Day)
    rest_day_hourly_rate_overtime = float(hourly_rate) * float(OT_Rest_Day)

    daily_rate = round(daily_rate, 2)
    hourly_rate = round(hourly_rate, 2)
    min_rate = round(min_rate, 2) 

    invalid_time = datetime.datetime.strptime('0:00', '%H:%M')
    time_in = datetime.datetime.strptime(data['TimeIn'], '%H:%M')
    time_out = datetime.datetime.strptime(data['TimeOut'], '%H:%M')

    # Constant variables for payroll breakdown
    OT_HOURS = 0
    LATE_MIN = 0
    UNDERTIME_MIN = 0
    LATE_UNDERTIME_AMOUNT = 0
    ABSENCES = 0
    ABSENCES_AMOUNT = 0

    DEDUCTIONS = 0
    OT_PAY = 0
    HOLIDAY_PAY = 0 

    if data['Day'] in REST_DAYS:
        # During rest days
        if time_in != invalid_time and time_out != invalid_time:
            # When time in and time out is invalid
            if time_out > time_in:
                if data['OTCategory'] != OVERTIME_CATEGORY[0] and data['Holiday'] != HOLIDAY_CATEGORY[0]:
                    overtime = float(data['Overtime']) 
                    if data['Holiday'] == HOLIDAY_CATEGORY[1]:
                        # Regular Holiday
                        HOLIDAY_PAY = HOLIDAY_PAY + (float(daily_rate) * Regular_Holiday_Pay) 
                        if overtime > 0.00:
                            if data['OTCategory'] == OVERTIME_CATEGORY[1]:
                                # Regular OT
                                daily_rate = (float(daily_rate) * Regular_Holiday_Pay / 8.00 * OT_Regular_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate
                            if data['OTCategory'] == OVERTIME_CATEGORY[2]:
                                # Regular OT  
                                daily_rate = (float(daily_rate) * 2.00 / 8.00 * OT_Regular_Day) * float(overtime)
                                # daily_rate = float(daily_rate * Regular_Holiday_Pay / 8.00 * OT_Regular_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate
                            
                    elif data['Holiday'] == HOLIDAY_CATEGORY[2]:
                        HOLIDAY_PAY = HOLIDAY_PAY + (float(daily_rate) * Special_Holiday_Pay)
                        if overtime > 0.00:
                            if data['OTCategory'] == OVERTIME_CATEGORY[1]:
                                # Regular OT
                                daily_rate = (float(daily_rate) * Special_Holiday_Pay / 8.00 * OT_Regular_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate
                            if data['OTCategory'] == OVERTIME_CATEGORY[2]:
                                # Rest day OT
                                daily_rate = (float(daily_rate) * Special_Holiday_Pay / 8.00 * OT_Rest_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate

                elif data['OTCategory'] != OVERTIME_CATEGORY[0]:  
                    # If category is excluded regular days 
                    overtime = float(data['Overtime'])
                    if overtime > 0.00: 
                        if data['OTCategory'] == OVERTIME_CATEGORY[1]:
                            t_diff = int((time_out - time_in).total_seconds() / 60.0)
                            t_diff_h = round(float(t_diff/60),2)   
                            total_overtime = round(t_diff_h * float(regular_hourly_rate_overtime),2) 
                        if  data['OTCategory'] == OVERTIME_CATEGORY[2]:
                            t_diff = int((time_out - time_in).total_seconds() / 60.0)
                            t_diff_h = round(float(t_diff/60),2)   
                            total_overtime = round(t_diff_h * float(rest_day_hourly_rate_overtime),2)  
    elif data['Day'] in WORKING_DAYS:
        if data['Itinerary'] or data['Leave']:
            # If the day was Itineraried or leaved safe!
            pass 
        else:
           
            if time_in != invalid_time and time_out != invalid_time:
                # If present
                # If time is valid
                if data['Late']:
                    # Calculate Late
                    l = 0
                    late = data['Late'] 
                    LATE_MIN += float(late)
                    deduction = float(late) * float(min_rate)
                    if deduction > daily_rate:
                        l = float(daily_rate)
                    else: 
                        l = l + deduction     
                    LATE_UNDERTIME_AMOUNT = LATE_UNDERTIME_AMOUNT + l
                    DEDUCTIONS = DEDUCTIONS + l  
                if data['Undertime']:
                    # Calculating Undertime
                    u = 0
                    undertime = data['Undertime']  
                    UNDERTIME_MIN += float(undertime)
                    deduction = float(undertime) * float(min_rate)
                    if deduction > daily_rate:
                        u = float(daily_rate)
                    else: 
                        u = u + deduction 
                    LATE_UNDERTIME_AMOUNT = LATE_UNDERTIME_AMOUNT + u
                    DEDUCTIONS = DEDUCTIONS + u 
                if data['OTCategory'] != OVERTIME_CATEGORY[0] and data['Holiday'] != HOLIDAY_CATEGORY[0]:
                    overtime = float(data['Overtime']) 
                    OT_HOURS += overtime
                    if data['Holiday'] == HOLIDAY_CATEGORY[1]:
                        # Regular Holiday
                        HOLIDAY_PAY = HOLIDAY_PAY + (float(daily_rate) * Regular_Holiday_Pay) 
                        if overtime > 0.00:
                            if data['OTCategory'] == OVERTIME_CATEGORY[1]:
                                # Regular OT
                                daily_rate = (float(daily_rate) * Regular_Holiday_Pay / 8.00 * OT_Regular_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate
                            if data['OTCategory'] == OVERTIME_CATEGORY[2]:
                                # Regular OT  
                                daily_rate = (float(daily_rate) * 2.00 / 8.00 * OT_Regular_Day) * float(overtime)
                                # daily_rate = float(daily_rate * Regular_Holiday_Pay / 8.00 * OT_Regular_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate
                            
                    elif data['Holiday'] == HOLIDAY_CATEGORY[2]:
                        HOLIDAY_PAY = HOLIDAY_PAY + (float(daily_rate) * Special_Holiday_Pay)
                        if overtime > 0.00:
                            if data['OTCategory'] == OVERTIME_CATEGORY[1]:
                                # Regular OT
                                daily_rate = (float(daily_rate) * Special_Holiday_Pay / 8.00 * OT_Regular_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate
                            if data['OTCategory'] == OVERTIME_CATEGORY[2]:
                                # Rest day OT
                                daily_rate = (float(daily_rate) * Special_Holiday_Pay / 8.00 * OT_Rest_Day) * float(overtime) 
                                OT_PAY = OT_PAY + daily_rate

                elif data['OTCategory'] != OVERTIME_CATEGORY[0]: 
                    # When you have overtime 
                    overtime = float(data['Overtime']) 
                    if overtime > 0.00:
                        OT_HOURS += overtime
                        if data['OTCategory'] == OVERTIME_CATEGORY[1]:
                            # Regular OT
                            OT_PAY = OT_PAY + (float(overtime) * regular_hourly_rate_overtime) 
                        if data['OTCategory'] == OVERTIME_CATEGORY[2]:
                            # Rest day OT
                            OT_PAY = OT_PAY + (float(overtime) * rest_day_hourly_rate_overtime) 
                    else:
                        total_overtime = total_overtime + 0.00

                elif data['Holiday'] != HOLIDAY_CATEGORY[0]:
                    if data['Holiday'] == HOLIDAY_CATEGORY[1]:
                        # Regular Holiday
                        HOLIDAY_PAY = HOLIDAY_PAY + (float(daily_rate) * Regular_Holiday_Pay)  
                    elif data['Holiday'] == HOLIDAY_CATEGORY[2]:
                        HOLIDAY_PAY = HOLIDAY_PAY + (float(daily_rate) * Special_Holiday_Pay) 

             # Working Day
            else:
                # If no IN or OUt
                if data['Holiday'] != "No Holiday":
                    # If Holiday
                    # Do nothing!
                    pass
                else:
                    # Your Absent
                    ABSENCES += 1
                    ABSENCES_AMOUNT = ABSENCES_AMOUNT + float(daily_rate)  
                    DEDUCTIONS = DEDUCTIONS + ABSENCES_AMOUNT
    
    PAYROLL_INFO.update({"OT_HOURS": round(float(OT_HOURS),2)})
    PAYROLL_INFO.update({"LATE_MIN": round(float(LATE_MIN),2)})
    PAYROLL_INFO.update({"UNDERTIME_MIN": round(float(UNDERTIME_MIN),2)})
    PAYROLL_INFO.update({"LATE_UNDERTIME_AMOUNT": round(float(LATE_UNDERTIME_AMOUNT),2)})
    PAYROLL_INFO.update({"ABSENCES": round(float(ABSENCES),2)})
    PAYROLL_INFO.update({"ABSENCES_AMOUNT": round(float(ABSENCES_AMOUNT),2)})
    PAYROLL_INFO.update({"DEDUCTIONS": round(float(DEDUCTIONS),2)})
    PAYROLL_INFO.update({"OT_PAY": round(float(OT_PAY),2)})
    PAYROLL_INFO.update({"HOLIDAY_PAY": round(float(HOLIDAY_PAY),2)})
    # PAYROLL_INFO.update({"total_overtime": round(float(total_overtime),2)})
    # PAYROLL_INFO.update({"total_absences": round(float(total_late_absences),2)})
    return PAYROLL_INFO

@login_required
def employee_create_payroll(request, key, id): 
    template_name = "payroll/employee_create_payroll.html" 
    user = get_object_or_404(User, username=request.user.username)      
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        pid = decrypt_key(key)
        employee = get_object_or_404(PersonalInfo, id=pid) 
        company_details = get_object_or_404(User, username=employee.fk_user.username) 
        employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee)
        
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

        #https://www.programiz.com/python-programming/datetime/current-datetime
        today = date.today()
        date_today = today.strftime("%m/%d/%Y") 
        attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period=cutoff)).order_by("id").distinct()
        
        #check if payroll exists
        emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
        
        
        total_deduction = 0.00

        T_OT_HOURS = 0
        T_LATE_MIN = 0
        T_UNDERTIME_MIN = 0
        T_LATE_UNDERTIME_AMOUNT = 0
        T_ABSENCES = 0 
        T_ABSENCES_AMOUNT = 0
        T_DEDUCTIONS = 0
        T_OT_PAY = 0
        T_HOLIDAY_PAY = 0

        #print(attendance,attendance.count())
        for data in attendance:
            
            # [on_true] if [expression] else [on_false] 
            dTIn = data.time_in if data.time_in  else '0:00'
            dTOut = data.time_out if data.time_out  else '0:00'
            output = 'day: {day} | date: {date} | time-in: {timein} | time-out: {timeout} | late: {late} | undertime: {undertime} | overtime: {overtime} | itinerary: {itinerary} | leave: {leave} | category: {category} | holiday: {holiday}'.format(day=data.days_of_week, date=data.date, timein=dTIn, timeout=dTOut, late=data.late, undertime=data.undertime, overtime=data.overtime, itinerary=data.has_itenerary, leave=data.has_leave, category=data.overtime_category, holiday=data.holiday)
            # print('->',output)
            attendance_data = {
                "Day": data.days_of_week,
                "TimeIn": dTIn,
                "TimeOut": dTOut, 
                "Late":data.late,
                "Undertime": data.undertime,
                "Overtime": data.overtime, 
                "Itinerary": data.has_itenerary,
                "Leave": data.has_leave,
                "OTCategory":data.overtime_category,
                "Holiday": data.holiday,
                "employee_salary":employee_salary.amount,
                "employee_allowance": employee_salary.allowance,
            } 
            #calculation 
            output = day_shift_payroll_computation_v1_2(attendance_data) 
            # total_lates_absences = total_lates_absences + output['total_absences'] 
            # total_overtime = total_overtime + output['total_overtime']
            T_OT_HOURS += output['OT_HOURS'] 
            T_LATE_MIN += output['LATE_MIN'] 
            T_UNDERTIME_MIN += output['UNDERTIME_MIN'] 
            T_LATE_UNDERTIME_AMOUNT += output['LATE_UNDERTIME_AMOUNT'] 
            T_ABSENCES += output['ABSENCES']  
            T_ABSENCES_AMOUNT += output['ABSENCES_AMOUNT'] 
            T_DEDUCTIONS += output['DEDUCTIONS']  
            T_OT_PAY += output['OT_PAY'] 
            T_HOLIDAY_PAY += output['HOLIDAY_PAY'] 

        print("OT_HOURS:",T_OT_HOURS)
        print("LATE_MIN:",T_LATE_MIN)
        print("UNDERTIME_MIN:",T_UNDERTIME_MIN)
        print("LATE_UNDERTIME_AMOUNT:",T_LATE_UNDERTIME_AMOUNT)
        print("ABSENCES:",T_ABSENCES)
        print("ABSENCES_AMOUNT:",T_ABSENCES_AMOUNT)
        print("T_DEDUCTIONS:",T_DEDUCTIONS)
        print("OT_PAY:",T_OT_PAY)
        print("HOLIDAY_PAY:",T_HOLIDAY_PAY) 

        # total_sss_contribution = round(sss_contribution(float(employee_salary.amount)),2)
        # total_philhealth_contribution = round(phlhealth_contribution(float(employee_salary.amount)),2)

        # total_deductions = total_lates_absences + total_sss_contribution + total_philhealth_contribution
        # total_gross_pay = float(employee_salary.amount/2) + float(employee_salary.allowance/2)  + total_overtime
        # total_net_pay = total_gross_pay - total_deductions
    

        # print('TOTAL OT: {ot} | TOTAL DEDUCT: {deduct}'.format(ot=total_overtime,deduct=total_deductions)) 
    
        total_sss_contribution = 0
        total_philhealth_contribution = 0
        
        total_deduction = T_DEDUCTIONS
        total_gross_pay = float(employee_salary.amount/2) + float(employee_salary.allowance/2)  + T_OT_PAY + T_HOLIDAY_PAY
        total_net_pay = total_gross_pay - total_deduction
        
        if request.method == 'GET':
                #check if there is an existing record  
                if emp_payroll_count > 0: 
                    #raise Http404()
                    return HttpResponseRedirect(reverse_lazy('application:employee_edit_payroll', kwargs={'key':employee.key_id,'id': cutoff.id}))
            
        elif request.method == 'POST': 
            _basicPay = request.POST.get('_basicPay',"0") if request.POST.get('_basicPay',"0")  != "" else "0"
            _allowance = request.POST.get('_allowance',"0") if request.POST.get('_allowance',"0")  != "" else "0"

            _overtimeHours = request.POST.get('_overtimeHours',"0") if request.POST.get('_overtimeHours',"0")  != "" else "0"
            _overtimePay = request.POST.get('_overtimePay',"0") if request.POST.get('_overtimePay',"0")  != "" else "0"
            _holidayPay = request.POST.get('_holidayPay',"0") if request.POST.get('_holidayPay',"0")  != "" else "0"

            # _sundaySpecialHoliday = request.POST.get('_sundaySpecialHoliday',"0") if request.POST.get('_sundaySpecialHoliday',"0")  != "" else "0"
            
            _salaryCashAdvance = request.POST.get('_salaryCashAdvance',"0") if request.POST.get('_salaryCashAdvance',"0")  != "" else "0"
            _grossPay = request.POST.get('_grossPay',"0") if request.POST.get('_grossPay',"0")  != "" else "0"
            _netPay = request.POST.get('_netPay',"0") if request.POST.get('_netPay',"0")  != "" else "0" 
            
            _lateMin = request.POST.get('_lateMin',"0") if request.POST.get('_lateMin',"0")  != "" else "0"
            _undertimeMin = request.POST.get('_undertimeMin',"0") if request.POST.get('_undertimeMin',"0")  != "" else "0"
            _lateUndertimeMinAmount = request.POST.get('_lateUndertimeMinAmount',"0") if request.POST.get('_lateUndertimeMinAmount',"0")  != "" else "0"
            _absences = request.POST.get('_absences',"0") if request.POST.get('_absences',"0")  != "" else "0"
            _absencesAmount = request.POST.get('_absencesAmount',"0") if request.POST.get('_absencesAmount',"0")  != "" else "0"

            _philhealContribution = request.POST.get('_philhealContribution',"0") if request.POST.get('_philhealContribution',"0")  != "" else "0"
            _pagibigContribution = request.POST.get('_pagibigContribution',"0") if request.POST.get('_pagibigContribution',"0")  != "" else "0"
            _sssPremius = request.POST.get('_sssPremius',"0") if request.POST.get('_sssPremius',"0")  != "" else "0"
            _withholdingTax = request.POST.get('_withholdingTax',"0") if request.POST.get('_withholdingTax',"0")  != "" else "0"
            _pagibigLoan = request.POST.get('_pagibigLoan',"0") if request.POST.get('_pagibigLoan',"0")  != "" else "0"
            _deductionSalaryCashAdvance = request.POST.get('_deductionSalaryCashAdvance',"0") if request.POST.get('_deductionSalaryCashAdvance',"0")  != "" else "0"
            _totalDeduction = request.POST.get('_totalDeduction',"0") if request.POST.get('_totalDeduction',"0")  != "" else "0"
        
            
            new, existing = EmployeePayroll.objects.update_or_create(employee_fk=employee, payroll_cutoff_period=cutoff, 
            defaults= { 
                'basic_pay': round(float(employee_salary.amount/2), 2),
                'allowance': round(float(employee_salary.allowance/2), 2),
                'ot_hours': float(_overtimeHours), 
                'ot_pay': round(float(_overtimePay), 2), 
                'holiday_pay': round(float(_holidayPay), 2), 
                'salary_or_cash_advance': _salaryCashAdvance,
                'gross_pay': _grossPay,

                'late_min': round(float(_lateMin), 2),
                'undertime_min': round(float(_undertimeMin), 2),
                'late_undertime_min_amount': round(float(_lateUndertimeMinAmount), 2),
                'absences': round(float(_absences), 2),
                'absences_amount': round(float(_absencesAmount), 2),

                'sss_premiums': _sssPremius,
                'philhealth_contribution': _philhealContribution,
                'pagibig_contribution': _pagibigContribution,
                'withholding_tax': _withholdingTax,
                'pagibig_loan': _pagibigLoan,
                'deducted_salary_cash_advance': _deductionSalaryCashAdvance,
                'total_deduction': _totalDeduction,
                'net_pay': float(_netPay), 
                })
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_payroll_page', kwargs={'id':cutoff.id}))
            Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your payroll for cut-off {cutoffperiod} was created!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])
            return HttpResponseRedirect(reverse_lazy('application:employee_payroll_manage_employee_list', kwargs={'id': cutoff.id}))
        
        context = {
            'user':user, 
            'cutoff':cutoff,
            'employee':employee,
            'date_today':date_today,
            'company_details': company_details,
            'attendance':attendance,
            'employee_salary':employee_salary,

            'basic_pay': round(float(employee_salary.amount/2), 2),
            'allowance' : round(float(employee_salary.allowance/2), 2),
            # 'legal_holiday': round(legal_holiday, 2),
            # 'sunday_special_holiday': round(sunday_spec_holiday, 2),

            'T_OT_HOURS':T_OT_HOURS,
            'T_LATE_MIN':T_LATE_MIN,
            'T_UNDERTIME_MIN':T_UNDERTIME_MIN,
            'T_LATE_UNDERTIME_AMOUNT': round(T_LATE_UNDERTIME_AMOUNT,2),
            'T_ABSENCES':T_ABSENCES,
            'T_ABSENCES_AMOUNT': round(T_ABSENCES_AMOUNT,2),
            'T_DEDUCTIONS':round(T_DEDUCTIONS,2),
            'T_OT_PAY':round(T_OT_PAY,2),
            'T_HOLIDAY_PAY':round(T_HOLIDAY_PAY,2),

            # 'sss_contribution': total_sss_contribution,
            # 'philhealth_contribution': total_philhealth_contribution,
            # 'pagibig_contribution': 0,
            'gross_pay': round(total_gross_pay, 2), 
            'net_pay': round(total_net_pay, 2),

            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 		
		  
@login_required
def employee_edit_payroll(request, key, id): 
    template_name = "payroll/employee_edit_payroll.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:  
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        pid = decrypt_key(key)
        employee = get_object_or_404(PersonalInfo, id=pid)  
        employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee)
        company_details = get_object_or_404(User, username=employee.fk_user.username)
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

        today = date.today()
        date_today = today.strftime("%m/%d/%Y") 

        attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period=cutoff)).order_by("-id").distinct()
        
        #check if payroll exists
        emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
        

        if request.method == 'GET':
                
            if emp_payroll_count <= 0: 
                return HttpResponseRedirect(reverse_lazy('application:employee_create_payroll', kwargs={'key':employee.key_id,'id': cutoff.id}))
            else: 
                emp_payroll = get_object_or_404(EmployeePayroll, employee_fk=employee, payroll_cutoff_period=cutoff)
                forms = EmployeePayrollForm(request.GET or None, instance=emp_payroll) 

        elif request.method == 'POST':
            emp_payroll = get_object_or_404(EmployeePayroll, employee_fk=employee, payroll_cutoff_period=cutoff)
            forms = EmployeePayrollForm(request.POST or None, instance=emp_payroll) 
            if forms.is_valid():
                instance = forms.save(commit=False)
                instance.save() 
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_payroll_page', kwargs={'id':cutoff.id}))
                Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your payroll for cut-off {cutoffperiod} was updated!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])               
                return HttpResponseRedirect(reverse_lazy('application:employee_payroll_manage_employee_list', kwargs={'id': cutoff.id}))
            else:
                print("Form Error!")
        context = {
            'user':user, 
            'cutoff':cutoff,
            'employee':employee,
            'date_today':date_today,
            'employee_salary': employee_salary,
            'attendance': attendance,
            'company_details':company_details,
            'forms': forms, 
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)

    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

@login_required
def employee_print_payroll(request, key, id):
    template_name = "payroll/employee_print_payroll.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        pid = decrypt_key(key)
        employee = get_object_or_404(PersonalInfo, id=pid)  
        employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee)
        company_details = get_object_or_404(User, username=employee.fk_user.username) 

        today = date.today()
        date_today = today.strftime("%m/%d/%Y") 

        attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period=cutoff)).order_by("-id").distinct()
        
        #check if payroll exists
        emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
        
    
        if request.method == 'GET': 
            emp_payroll = get_object_or_404(EmployeePayroll, employee_fk=employee, payroll_cutoff_period=cutoff)         
        context = {
            'user':user, 
            'cutoff':cutoff,
            'employee':employee,
            'date_today':date_today,
            'employee_salary': employee_salary,
            'attendance': attendance,
            'company_details':company_details,  
            'emp_payroll_count':emp_payroll,
        }

        return render(request, template_name, context) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_delete_payroll(request, key, id):
    data = dict()
    template_name = "payroll/employee_delete_payroll.html"
    user = get_object_or_404(User, username=request.user.username)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    pid = decrypt_key(key)
    employee = get_object_or_404(PersonalInfo, id=pid)  
    emp_payroll = get_object_or_404(EmployeePayroll, employee_fk=employee, payroll_cutoff_period=cutoff)
    link = HttpResponseRedirect(reverse_lazy('application:employee_payroll_manage_employee_list', kwargs={'id': cutoff.id}))
    if request.is_ajax():
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            emp_payroll.delete()
            data['form_is_valid'] = True
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_payroll_page', kwargs={'id':cutoff.id}))
            Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your payroll for cut-off {cutoffperiod} was deleted!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[1],level=level_list[1])               
         
        context = {
            'user':user, 
            'cutoff':cutoff,
            'employee':employee,
            'emp_payroll': emp_payroll, 
        }
        data['link'] = link.url
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_manage_leaves(request):
    template_name = "leaves/employee_leaves_page.html"
    user = get_object_or_404(User, username=request.user.username)  
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:  
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        leaves = EmployeeLeaves.objects.all().filter().order_by('-id').distinct() 
        if request.method == 'GET':
            try:  
                employee = PersonalInfo.objects.get(fk_user=user) 
                context = {
                    'user':user,
                    'employee': employee,
                    'leaves': leaves,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }
                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
  
@login_required
def employee_view_employee_leaves(request, id):
    template_name="leaves/employee_view_leave_page.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        leave = get_object_or_404(EmployeeLeaves, id=id)
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

        employee_list = PersonalInfo.objects.all().order_by('-id').distinct()

        status_list = [
            'Pending',
            'Disapproved',
            'Approved',
        ]


        employee_sender = get_object_or_404(EmployeeLeaves, id=id)  
        eid = employee_sender.employee_leave_fk.fk_user.company_to_user.id
        company_employee = CompanyInfo.objects.get(id=eid)
        if request.method == 'GET':    
            is_vl = True if employee_sender.classification_of_leave == "Vacation Leave" else False
            is_sl = True if employee_sender.classification_of_leave == "Sick Leave" else False 
        elif request.method == 'POST': 
            vl_credits = company_employee.vacation_leave_credits
            sl_credits = company_employee.sick_leave_credits
            # company_employee.vacation_leave_credits = 7
            # company_employee.save()

            _leaveCredits = request.POST.get('_leaveCredits', "0")
            _lessThisApplication = request.POST.get('_lessThisApplication', "0")
            _balanceAsOfThisDate = request.POST.get('_balanceAsOfThisDate', "0")
            _status = request.POST['statusList']
            _notedBy = request.POST['notedByList']
            _checkedBy = request.POST['checkedByList']
            _approvedBy = request.POST['approvedByList']

            #updated
            _leaveCredits = 0
            _lessThisApplication = 0
            _balanceAsOfThisDate = 0 

            _remarks = ""
            _has_payment = False

            if _status.casefold() == 'Approved'.casefold(): 
                if employee_sender.status.casefold() != 'Approved'.casefold(): 
                    if employee_sender.classification_of_leave == "Vacation Leave" or employee_sender.classification_of_leave == "Bereavement Leave":
                        if vl_credits > 0:
                            no_days = employee_sender.no_days
                            balance = vl_credits - no_days
                            company_employee.vacation_leave_credits = balance
                            company_employee.save()
                            _leaveCredits = vl_credits
                            _lessThisApplication = no_days
                            _balanceAsOfThisDate = balance 
                            _has_payment = True 
                            if balance < 0:
                                balance = balance * -1 
                                _remarks = 'The {balance} are days not classified as leave with PAYMENT'.format(balance=balance)
                    elif employee_sender.classification_of_leave == "Sick Leave": 
                        if sl_credits > 0:
                            no_days = employee_sender.no_days
                            balance = sl_credits - no_days
                            company_employee.sick_leave_credits = balance
                            company_employee.save()
                            _leaveCredits = vl_credits
                            _lessThisApplication = no_days
                            _balanceAsOfThisDate = balance 
                            _has_payment = True 
                            # add sentence!
                            if balance < 0:
                                balance = balance * -1
                                _remarks = 'The {balance} are days not classified as leave with PAYMENT'.format(balance=balance)
            elif _status.casefold() == 'Disapproved'.casefold():  
                if employee_sender.status.casefold() != 'Disapproved'.casefold(): 
                    if employee_sender.classification_of_leave == "Vacation Leave" or employee_sender.classification_of_leave == "Bereavement Leave":
                        if vl_credits < 7:
                            no_days = employee_sender.no_days
                            balance = vl_credits + no_days
                            company_employee.vacation_leave_credits = balance
                            company_employee.save() 
                    elif employee_sender.classification_of_leave == "Sick Leave":
                        if sl_credits < 7:
                            no_days = employee_sender.no_days
                            balance = sl_credits + no_days
                            company_employee.sick_leave_credits = balance
                            company_employee.save()  
                    
            new, existing = EmployeeLeaves.objects.update_or_create(id=id, defaults={
                'status': _status,
                'has_payment': _has_payment,
                'remarks':_remarks,
                'leave_credits': _leaveCredits,
                'less_this_application': _lessThisApplication,
                'balance_as_of_this_date': _balanceAsOfThisDate, 
                'noted_by': _notedBy,
                'checked_by': _checkedBy,
                'approved_by': _approvedBy,

            })
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_leave_form', kwargs={'id':leave.id}))
            Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your leave form was updated by the administrator!",category=category_list[10],level=level_list[1])               
            return HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
        context = {
            'user':user,
            'company_employee':company_employee,
            'is_vl':is_vl,
            'is_sl':is_sl,
            'employee': employee,
            'leave': leave, 
            'employee_list': employee_list,
            'status_list': status_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_delete_leave_form(request, id):
    data = dict()
    template_name = "leaves/employee_delete_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    leave = get_object_or_404(EmployeeLeaves, id=id) 
    if request.is_ajax():
        if user.is_active and user.is_staff and user.is_superuser:
            if request.method == 'GET':
                
                context = {
                'user': user,  
                'employee': employee,   
                'leave': leave,          
                } 
                data['html_form'] = render_to_string(template_name, context, request)
            elif request.method == 'POST':
                eid = leave.employee_leave_fk.fk_user.company_to_user.id
                company_employee = CompanyInfo.objects.get(id=eid)
                vl_credits = company_employee.vacation_leave_credits
                sl_credits = company_employee.sick_leave_credits
                if leave.classification_of_leave == "Vacation Leave" or leave.classification_of_leave == "Bereavement Leave":
                    if vl_credits < 7:
                        no_days = leave.no_days
                        balance = vl_credits + no_days
                        company_employee.vacation_leave_credits = balance
                        company_employee.save() 
                elif leave.classification_of_leave == "Sick Leave":
                    if sl_credits < 7:
                        no_days = leave.no_days
                        balance = sl_credits + no_days
                        company_employee.sick_leave_credits = balance
                        company_employee.save()  


                leave.delete()
                data['form_is_valid'] = True 
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_leaves_page'))  
                Notifications.objects.create(sender=user,recipient=leave.employee_leave_fk.fk_user,url=url.url,message="Leave form from {sender} has been deleted!".format(sender=user),category=category_list[11],level=level_list[1])
                
            return JsonResponse(data)
        else:
            raise Http404()
    else:
        raise Http404()

@login_required
def employee_manage_itenerary(request):
    template_name = "iteneraries/employee_manage_iteneraries_page.html"
    user = get_object_or_404(User, username=request.user.username)  
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        itenerary_list = EmployeeItenerary.objects.all().order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count() 
        if request.method == 'GET':
            try:  
                employee = PersonalInfo.objects.get(fk_user=user) 
                context = {
                    'user':user,
                    'employee': employee, 
                    'itenerary_list': itenerary_list,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }
                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))        
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_view_itenerary(request, id):
    template_name = "iteneraries/employee_view_iteneraries_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        itenerary = get_object_or_404(EmployeeItenerary, id=id)
        employee_list = PersonalInfo.objects.all().order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()

    
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            _notedBy = request.POST['notedByList']
            _checkedBy = request.POST['checkedByList']
            _approvedBy = request.POST['approvedByList']


            new, existing = EmployeeItenerary.objects.update_or_create(id=id, defaults={ 
                'noted_by': _notedBy,
                'checked_by': _checkedBy,
                'approved_by': _approvedBy,

            })
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_iteneraries_page', kwargs={'id':itenerary.id}))
            Notifications.objects.create(sender=user,recipient=itenerary.employee_itenerary_fk.fk_user,url=url.url,message="Your itinerary form was updated by the administrator!",category=category_list[10],level=level_list[1])               
            return HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
        context = {
            'user':user, 
            'itenerary': itenerary,
            'employee_list': employee_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_manage_concerns(request):
    template_name = "concerns/employee_manage_concern.html"
    user = get_object_or_404(User, username=request.user.username)  
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count() 
        if request.method == 'GET':
            try:  
                employee = PersonalInfo.objects.get(fk_user=user) 
                concern_list = Concerns.objects.all().filter(Q(sender=employee)).order_by('-id').distinct()
                context = {
                    'user':user,
                    'employee': employee, 
                    'concern_list': concern_list,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }
                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

@login_required
def employee_create_concern(request):
    data = dict()
    template_name = "concerns/employee_create_concern.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user) 
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None)   
            if form.is_valid():
                instance = form.save(commit=False)         
                instance.sender  = employee
                instance.save()
                #notification
                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="You had received new concern from {sender}".format(sender=employee),category=category_list[9],level=level_list[1])               
                #end notificaiton
                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')
                edit_url = reverse_lazy('application:employee_edit_concerns_page', kwargs={'id':instance.id})
                delete_url = reverse_lazy('application:employee_delete_concerns_page', kwargs={'id':instance.id})
                view_url = reverse_lazy('application:employee_view_concerns_page', kwargs={'id':instance.id})
                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))   

@login_required
def employee_edit_concern(request, id):
    data = dict()
    template_name = "concerns/employee_edit_concern.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id) 
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)         
                instance.sender  = employee
                instance.save()

                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="Concern from {sender} was updated".format(sender=employee),category=category_list[10],level=level_list[1])               

                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')
                edit_url = reverse_lazy('application:employee_edit_concerns_page', kwargs={'id':instance.id})
                delete_url = reverse_lazy('application:employee_delete_concerns_page', kwargs={'id':instance.id})
                view_url = reverse_lazy('application:employee_view_concerns_page', kwargs={'id':instance.id})
                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True

        context = {
            'form': form,
            'concern': concern,
            'user': user,  
            'employee': employee,  
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_view_concern(request, id):
    data = dict()
    template_name = "concerns/employee_view_concern.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id) 
        if request.method == 'GET':
            pass
       
        context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
        } 
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_delete_concern(request, id):
    data = dict()
    template_name = "concerns/employee_delete_concern.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id) 
        if request.method == 'GET':
            context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
            } 
            data['html_form'] = render_to_string(template_name, context, request)
        elif request.method == 'POST': 
            if concern.receiver.fk_user.is_superuser:
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
            else:
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page')) 
            Notifications.objects.create(sender=user,recipient=concern.receiver.fk_user,url=url.url,message="Your concern from {sender}".format(sender=concern.sender),category=category_list[11],level=level_list[1])               
            concern.delete()
            data['form_is_valid'] = True 
            
        return JsonResponse(data)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
 
@login_required
def employee_manage_inbox_concern(request):
    template_name = "concerns/employee_manage_inbox_concern.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count() 
        if request.method == 'GET':
            try:  
                employee = PersonalInfo.objects.get(fk_user=user) 
                concern_list = Concerns.objects.all().filter(Q(receiver=employee)).order_by('-id').distinct()
                context = {
                    'user': user,  
                    'employee': employee, 
                    'concern_list': concern_list,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }

                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))         
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

@login_required
def employee_reply_concern(request, id):
    data = dict()
    template_name = "concerns/employee_reply_concern.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id) 
        if request.method == 'GET':
            form = ConcernsReplyEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsReplyEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)   
                instance.save()  
                if instance.sender.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.sender.fk_user,url=url.url,message="Your concern has been repilied by {receiver}".format(receiver=concern.receiver),category=category_list[8],level=level_list[1])               


                data['form_is_valid'] = True

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
            return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

@login_required
def employee_payroll_base_salary_settings(request, key):
    template_name = "employees/employee_salary_settings.html"
    current_user = get_object_or_404(User, username=request.user.username)
    data = dict()
    pid = decrypt_key(key)
    employee = get_object_or_404(PersonalInfo, id=pid) 
    
    count = EmployeeSalary.objects.all().filter(employee_salary_fk=employee).distinct()
 
    if count: 
        EmployeeSalaryInlineFomrset = inlineformset_factory(PersonalInfo, EmployeeSalary, form=EmployeeSalaryForm, extra=0 , can_delete=False)
    else:
        EmployeeSalaryInlineFomrset = inlineformset_factory(PersonalInfo, EmployeeSalary, form=EmployeeSalaryForm, extra=1 , can_delete=False)

    if request.is_ajax():
        if request.method == 'GET': 
            formset = EmployeeSalaryInlineFomrset(request.GET or None, instance=employee)
        elif request.method == 'POST':
            formset = EmployeeSalaryInlineFomrset(request.POST or None, instance=employee)
            if formset.is_valid():
                myform = formset.save(commit=False)
                for mf in myform:
                    mf.employee_salary_fk = employee
                    mf.save()
                data['form_is_valid'] = True
                Notifications.objects.create(sender=current_user,recipient=employee.fk_user,message="Your payroll settings was updated!",category=category_list[3],level=level_list[1])
            else:
                messages.error(request, "Form Error")


        context = {
            'employee': employee,
            'formset': formset,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def upload_attendance(request):
    data = dict()
    template_name='attendance/upload_attendance.html'

    if request.is_ajax(): 
        if request.method == 'GET':
            form = CutOffPeriodForm(request.GET or None, request.FILES or None)
        elif request.method == 'POST':
            form = CutOffPeriodForm(request.POST or None, request.FILES or None) 

            if form.is_valid():
                try:
                   
                    if 'attendance_file' in request.FILES:
                        att_file = request.FILES['attendance_file']
                        print(att_file.name) 
                        #https://code.djangoproject.com/ticket/26641
                        # employees = PersonalInfo.objects.values_list('id', flat=True).order_by('id')
                        #employees = PersonalInfo.objects.annotate(name=Concat('first_name', Value(' '),'middle_name', Value(' '),'last_name'),).filter(Q(name__icontains='lloyds Salazar Garcia'))
                        
                        # .query to raw sql query
                        #print(employees.query) 
                        #data['form_is_valid'] = True
                        read_attendance_file(att_file, request, form, data) 
                    else:
                        print('No File')                     
                except IntegrityError:
                    data['form_is_valid'] = False 
                    messages.error(request, 'Record Already Exists')
            else:
                data['form_is_valid'] = False
        context = {
            'form':form,
        }
        data['html_form'] =  render_to_string(template_name,context, request)
        return JsonResponse(data)
    else:
        raise Http404()

    # For Employees

@login_required
def employee_overtime_management_page(request):
    template_name = "overtime/employee_manage_overtime_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count() 
        if request.method == 'GET':
            try:  
                employee = PersonalInfo.objects.get(fk_user=user) 
                overtime_list = Overtime.objects.all().order_by('-id').distinct()
                context = {
                'user':user,
                'employee': employee, 
                'overtime_list': overtime_list,
                'notifications': notifications,
                'notifications_count': notifications_count,
                } 
                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))        
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

@login_required
def employee_view_overtime(request, id):
    template_name = "overtime/employee_view_overtime_page.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user) 
        overtime = get_object_or_404(Overtime, id=id) 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        employee_list = PersonalInfo.objects.all().order_by('-id').distinct()# for approval 
        overtime_list = Overtime.objects.all().filter(id=id).order_by('-id').distinct()
        overtime_sum = OvertimeDetails.objects.all().filter(overtime__in=overtime_list).aggregate(Sum('duration')).get('duration__sum')   
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            _notedBy = request.POST['notedByList']
            _checkedBy = request.POST['checkedByList']
            _approvedBy = request.POST['approvedByList']

            print(_notedBy)
            print(_approvedBy)

            new, existing = Overtime.objects.update_or_create(id=id, defaults={ 
                'noted_by': _notedBy,
                'checked_by': _checkedBy,
                'approved_by': _approvedBy,

            }) 
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_employee_view_overtime_form', kwargs={'id':overtime.id}))
            Notifications.objects.create(sender=user,recipient=overtime.employee_overtime.fk_user,url=url.url,message="Your overtime form was updated by the administrator!",category=category_list[10],level=level_list[1])               
            return HttpResponseRedirect(reverse_lazy('application:employee_overtime_management_page'))
        context = {
        'user':user,
        'employee': employee, 
        'employee_list': employee_list,
        'overtime':overtime,
        'overtime_list': overtime_list,
        'notifications': notifications,
        'notifications_count': notifications_count,
        'overtime_sum':overtime_sum,
        } 
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def employee_manage_roles_and_permission(request):
    template_name = "roles_and_permission/manage_employee_roles_and_permission.html"
    user = get_object_or_404(User, username=request.user.username)

    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

    roles_and_permission = RolesPermission.objects.all() 

    user = validate(user, is_managing_director=True)
    if user:
        users = User.objects.all().filter(Q(Q(is_staff=True)  & Q(is_active=True)) | Q(is_staff=True) & Q(is_active=True) & Q(is_superuser=True))

        list_of_users = []
        for u in users:
            try:
                if u.profile_to_user:
                    list_of_users.append(u.profile_to_user.first_name)
            except User.profile_to_user.RelatedObjectDoesNotExist:
                print("No child record!")
             
        list_of_users_data = {
            'list_of_users': list_of_users,
        }

        context = {
            'user': user,
            'users': users,
            'list_of_users_data': list_of_users_data,
            'notifications': notifications,
            'notifications_count': notifications_count,
            'roles_and_permission': roles_and_permission,
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

# Beta version for future testing
@login_required
def employee_view_roles_and_permission(request):
    template_name = "roles_and_permission/manage_roles_and_permission.html" 
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        users = User.objects.all().filter(Q(Q(is_staff=True)  & Q(is_active=True)) | Q(is_staff=True) & Q(is_active=True) & Q(is_superuser=True))
  
        list_of_users = []
        for u in users:
            try:
                if u.profile_to_user:
                    list_of_users.append(u.profile_to_user.first_name)
            except User.profile_to_user.RelatedObjectDoesNotExist:
                print("No child record!")
             
        list_of_users_data = {
            'list_of_users': list_of_users,
        }
        context = {
            'user':user,
            'list_of_users_data': list_of_users_data, 
        }
        return render(request, template_name, context) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

@login_required
def employee_create_roles_permission(request):
    data = dict()
    template_name = "roles_and_permission/create_new_role.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user) 
        roles_and_permission = RolesPermission.objects.all()
        list_of_employees = CompanyInfo.objects.filter(~Q(employee_ci_rp_fk_r__in=roles_and_permission)) 
    
        if request.method == 'GET':
            form = RolesPermissionForm(request.GET or None)
        elif request.method == 'POST':
            form = RolesPermissionForm(request.POST or None)  
            """
            json objects are surrounded by curly braces { }. They are written in key and value pairs. 
            json.loads() takes in a string and returns a json object. 
            json.dumps() takes in a json object and returns a string. 
            As you can see, json.dumps() and json.loads() are opposite of one another.
            """
             
            request_json = json.loads(request.body)
            foo = json.dumps(request_json, indent=4, sort_keys=True)
            # print(foo)
            
            _employee = request_json['employee']
            _role = request_json['role']
            _title = request_json['title']
            _immidiate_supervisor = request_json['immidiate_supervisor'] if request_json['immidiate_supervisor'] != None else 'None'
           
            # #employees = PersonalInfo.objects.annotate(name=Concat('first_name', Value(' '),'middle_name', Value(' '),'last_name'),).filter(Q(name__icontains='lloyds Salazar Garcia'))
            p_emp = PersonalInfo.objects.annotate(name=Concat('last_name', Value(' '),'first_name', Value(' '),'middle_name'),).get(Q(name__icontains=_employee))
            u_emp = get_object_or_404(User,profile_to_user=p_emp)
            c_emp = get_object_or_404(CompanyInfo, fk_company_user=u_emp) 
            new, existing = RolesPermission.objects.update_or_create(employee_ci_rp_fk=c_emp, defaults={
                'role':_role,
                'title':_title,
                'immidiate_head':_immidiate_supervisor
            })
 

            _title = str([t for t in _title ]).replace("'","").replace("[","").replace("]","")

            edit_role_url = reverse_lazy('application:employee_edit_roles_permission', kwargs={'id':new.id})
            delete_role_url = reverse_lazy('application:employee_delete_role_permission', kwargs={'id':new.id}) 

            data['record_data'] = {
                'employee': str(p_emp),
                'designation': c_emp.designation,
                'role': _role,
                'title': _title,
                'immidiate_supervisor':_immidiate_supervisor,
                'edit_role_url':edit_role_url,
                'delete_role_url': delete_role_url,
            }
            data['form_is_valid'] = True            
        context = {
            'form': form,
            'user': user,
            'employee': employee,
            'list_of_employees':list_of_employees,
        }        
        # roles
        list_emp = []
        roles_permission = RolesPermission.objects.all() 
        for rp in roles_permission:
            list_emp.append({"Role":str(rp.role),"Employee":str(rp.employee_ci_rp_fk.fk_company_user.profile_to_user)})

        bar = json.dumps(list_emp, indent=4, sort_keys=True)
        # print(bar) 
        data['list_of_employees'] = json.loads(bar)
        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:  
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
 
def remove_characters(remove_char, string):
    new_string = string
    for r_char in remove_char:       
        new_string = new_string.replace(r_char, "")   
    return new_string

# edit
@login_required
def employee_edit_roles_permission(request, id):
    data = dict()
    template_name = "roles_and_permission/edit_role.html"
    user = get_object_or_404(User, username=request.user.username)
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        roles_permission = get_object_or_404(RolesPermission, id=id) 
        if request.method == 'GET':
            _employee_name = roles_permission.employee_ci_rp_fk.fk_company_user.profile_to_user
            _role = roles_permission.role
            _title = roles_permission.title
            _immidiate_head = roles_permission.immidiate_head

            _title = list(remove_characters("[]'",_title).split(", ")) 

            rp_data = {
                '_employee_name': str(_employee_name),
                '_role': str(_role),
                '_title': _title,
                '_immidiate_head': str(_immidiate_head),
            }
            rp_data = json.dumps(rp_data, indent=4, sort_keys=True)   

            # roles
            list_emp = []
            roles_permission_list = RolesPermission.objects.all() 
            for rp in roles_permission_list:
                list_emp.append({"Role":str(rp.role),"Employee":str(rp.employee_ci_rp_fk.fk_company_user.profile_to_user)})

            list_emp = json.dumps(list_emp, indent=4, sort_keys=True) 
            data['list_emp'] = json.loads(list_emp) 
            data['rp_data'] = json.loads(rp_data)  

        elif request.method == 'POST':
            request_json = json.loads(request.body)
            foo = json.dumps(request_json, indent=4, sort_keys=True)
        

            _employee = request_json['employee']
            _role = request_json['role']
            _title = request_json['title']
            _immidiate_supervisor = request_json['immidiate_supervisor'] 

            p_emp = PersonalInfo.objects.annotate(name=Concat('last_name', Value(' '),'first_name', Value(' '),'middle_name'),).get(Q(name__icontains=_employee))
            u_emp = get_object_or_404(User,profile_to_user=p_emp)
            c_emp = get_object_or_404(CompanyInfo, fk_company_user=u_emp) 

            new, existing = RolesPermission.objects.update_or_create(employee_ci_rp_fk=c_emp, defaults={
                'role':_role,
                'title':_title,
                'immidiate_head':_immidiate_supervisor
            })

            _title = str([t for t in _title ]).replace("'","").replace("[","").replace("]","")

            rp_data = {
                '_employee_name': str(_employee),
                '_role': str(_role),
                '_title': _title,
                '_immidiate_supervisor': str(_immidiate_supervisor),
            }
            rp_data = json.dumps(rp_data, indent=4, sort_keys=True)   

            data['rp_data'] = json.loads(rp_data)  
            data['form_is_valid'] = True
        context = {
            'user':user,
            'roles_permission':roles_permission,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  
# delete
@login_required
def employee_delete_role_permission(request, id):
    data = dict()
    template_name = "roles_and_permission/delete_role.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        role_permission = get_object_or_404(RolesPermission,id=id) 
        if request.method == 'GET': 
            context = { 
            'user': user,   
            'role_permission': role_permission,
            } 
            data['html_form'] = render_to_string(template_name, context, request)
        elif request.method == 'POST':
            role_permission.delete()
            data['form_is_valid'] = True 

        return JsonResponse(data) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  

@login_required
def reports_page(request):
    template_name = "reports/reports_page.html"
    user = get_object_or_404(User, username=request.user.username) 
     
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count() 
        if request.method == 'GET':
         pass

        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count, 
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
  
@login_required
def admin_search_page(request):
    template_name = "admin_search_page.html"
    user = get_object_or_404(User, username=request.user.username)  
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user: 		 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()  
        if request.method == 'GET':
            search_term = request.GET.get('search_term')
            if search_term.strip():
                #when accessing child model from parent use related name!
                user_filter = User.objects.filter(
                    Q(Q(email__icontains=search_term) | 
                    #child personal info model User
                    Q(profile_to_user__first_name__icontains=search_term) | 
                    Q(profile_to_user__middle_name__icontains=search_term) | 
                    Q(profile_to_user__last_name__icontains=search_term) | 
                    Q(profile_to_user__dob__icontains=search_term) | 
                    Q(profile_to_user__gender__icontains=search_term) |
                    Q(profile_to_user__address__icontains=search_term) |
                    #child company info model User
                    Q(company_to_user__company_id__icontains=search_term) | 
                    Q(company_to_user__designation__icontains=search_term) |
                    Q(company_to_user__company_tin__icontains=search_term) |  
                    Q(company_to_user__personal_tin__icontains=search_term) | 
                    Q(company_to_user__sss_number__icontains=search_term) | 
                    Q(company_to_user__pagibig__icontains=search_term) | 
                    Q(company_to_user__philhealth__icontains=search_term) |
                    #mobile child info model User
                    Q(mobile_to_user__mobile_number__icontains=search_term) |
                    #telephone child info model User
                    Q(telephone_to_user__telephone_number__icontains=search_term) |
                    #skill child info model User
                    Q(skills_to_user__skills__icontains=search_term)) & Q(is_superuser=False)
                ).order_by('-id').distinct()#.values_list('email', flat=True)
                #personal_filter = PersonalInfo.objects.filter(Q(first_name__icontains=search_term)).order_by('-id').values_list('first_name','middle_name','last_name')[0][1]

                all_cutoff_list = CutOffPeriodInfo.objects.all().order_by('id').distinct()
                # Advance filter for every cut off there is multiple emp           
        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count,
            'search_term': search_term,
            'user_filter': user_filter, 
            'all_cutoff_list':all_cutoff_list,
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 
 
@login_required
def maintainance_page(request):
    template_name = "maintainance_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET':
            search_term = request.GET.get('search_term')
        if search_term.strip(): 
            print(search_term) 
        context = {
        'user':user, 
        'notifications': notifications,
        'notifications_count': notifications_count,
        'search_term': search_term,
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page')) 

@login_required
def employee_admin_error_page(request):
    template_name = "error/employee_admin_error_page.html"
    user = get_object_or_404(User, username=request.user.username)   
    if user.is_active and user.is_staff and user.is_superuser:
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count() 
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context) 
    else:
        raise Http404()
   
@login_required
def form_manager(request):
    template_name = "form_manager/form_manager.html"
    user = get_object_or_404(User, username=request.user.username)  
    user = validate(user, is_managing_director=True, is_human_resource=True) 
    if user:
        if request.method == 'GET':
            notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
            notifications_count = notifications.count()

            # get all record in the models
            leaves = list(EmployeeLeaves.objects.all().filter().order_by('-id').distinct() )
            itinerary_list = list(EmployeeItenerary.objects.all().order_by('-id').distinct())
            overtime_list = list(Overtime.objects.all().order_by('-id').distinct())
            all_forms = leaves + itinerary_list + overtime_list 

            for item in all_forms:
                if hasattr(item,'employee_leave_fk'):
                    print(item.employee_leave_fk)
                else:
                    continue
            context = {
                'user': user,   
                'all_forms': all_forms,
                'notifications': notifications,
                'notifications_count': notifications_count,
            }

            return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_admin_error_page'))  
# employee side-----------------------------------------------------------------------------
 
@login_required
def PayrallAttendanceEmployeeIndexPage(request):
    template_name = "employee_index_page.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user: 
        employee = PersonalInfo.objects.filter(fk_user=user).order_by('-id') 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        cutoffperiods = CutOffPeriodInfo.objects.all().distinct().count()
        total_leaves = EmployeeLeaves.objects.all().filter(Q(employee_leave_fk__in=employee)).distinct().count()  
        total_unapproved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Disapproved') & Q(employee_leave_fk__in=employee)).distinct().count()  
        total_approved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Approved')& Q(employee_leave_fk__in=employee)).distinct().count()  

        total_concern =  Concerns.objects.all().filter(Q(sender__in=employee)).distinct().count() 
        total_unreplied_concern = Concerns.objects.all().filter(Q(reply=None) & Q(sender__in=employee)).distinct().count() 
        
        total_unapproved_iteneraries = EmployeeItenerary.objects.all().filter(Q(~Q(checked_by=None) & ~Q(approved_by=None)) & Q(employee_itenerary_fk__in=employee)).distinct().count()  
        total_approved_iteneraries = EmployeeItenerary.objects.all().filter(Q(Q(checked_by=None) & Q(approved_by=None)) &  Q(employee_itenerary_fk__in=employee) ).distinct().count() 

        emp_payrolls = EmployeePayroll.objects.all().filter(Q(employee_fk__in=employee))
        year_list = []
        yearly_employee_salary_list = []
        yearly_emplpyee_deduction_list = []

        for emp in emp_payrolls:
            year = str(emp.payroll_cutoff_period)[:4]
            if not year in year_list:
                year_list.append(year) 

        for year in year_list:
            total_net_pay = 0
            total_deductions = 0
            cutoffperiod = CutOffPeriodInfo.objects.all().filter(Q(cut_off_period__icontains=year)) 
            for cutoff in cutoffperiod:
                emp_payroll = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff) & Q(employee_fk__in=employee)).aggregate(Sum('net_pay')).get('net_pay__sum')                   
                emp_deductions  = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff) & Q(employee_fk__in=employee)).aggregate(Sum('total_deduction')).get('total_deduction__sum') 
                
                if emp_payroll:
                    total_net_pay = total_net_pay + emp_payroll 
                    total_deductions = total_deductions + emp_deductions
            yearly_employee_salary_list.append({'year': year, 'value': int(round(total_net_pay,2))})
            yearly_emplpyee_deduction_list.append({'year': year, 'deduction': int(round(total_deductions,2)), 'wages': int(round(total_net_pay,2))})
        
        
        dumps_yearly_employee_salary_list = json.dumps(yearly_employee_salary_list)
        loads_yearly_employee_salary_list =json.loads(dumps_yearly_employee_salary_list)
        # print(dumps_yearly_employee_wages_list)
        
        dumps_yearly_employee_deduction_list = json.dumps(yearly_emplpyee_deduction_list)
        loads_yearly_employee_deduction_list =json.loads(dumps_yearly_employee_deduction_list)

        
        context = {
            'user': user,
            'cutoffperiods': cutoffperiods,
            'total_leaves':total_leaves,
            'total_unapproved_leaves': total_unapproved_leaves,
            'total_approved_leaves': total_approved_leaves,
            'total_concern': total_concern,
            'total_unreplied_concern': total_unreplied_concern,
            'total_unapproved_iteneraries':total_unapproved_iteneraries,
            'total_approved_iteneraries': total_approved_iteneraries,
            'loads_yearly_employee_salary_list': loads_yearly_employee_salary_list,
            'loads_yearly_employee_deduction_list':loads_yearly_employee_deduction_list, 
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 

@login_required
def employee_profile(request):
    template_name = "employee_side/employee_profile_page.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        last_time_logged = user.last_login
        last_time_joined = user.date_joined
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count() 
        if request.method == "GET": 
            record = PersonalInfo.objects.filter(fk_user=user)
            company = CompanyInfo.objects.filter(fk_company_user=user)
            profiles =  PersonalInfo.objects.all().filter(fk_user=user).distinct()
            company_details = CompanyInfo.objects.all().filter(fk_company_user=user).distinct()        
        elif request.method == "POST":
            pass
        context = {
            'user': user,
            'last_time_logged': last_time_logged,
            'last_time_joined': last_time_joined, 
            'record': record,
            'profiles': profiles,
            'company':company,
            'company_details':company_details,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
      

@login_required
def side_employee_cut_off_payroll_page(request):
    template_name="employee_side/employee_side_cutoff_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count() 
        cut_off_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()  
        if request.method == 'GET':
            pass

        context = {
            'user': user,
            'cut_off_list': cut_off_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
  
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
  

@login_required
def side_employee_cut_off_page(request, id):
    template_name = "employee_side/employee_side_cutoff_page.html"
    user = get_object_or_404(User, username=request.user.username)  
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        if request.method == 'GET':
            try:  
                employee = PersonalInfo.objects.get(fk_user=user) 
                attendance_list = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
                notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
                notifications_count = notifications.count()
                context = {
                    'user': user, 
                    'cutoff': cutoff,
                    'attendance_list': attendance_list,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }

                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
 

@login_required
def side_employee_view_payroll_page(request, id):
    template_name = "employee_side/employee_side_view_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET': 
            try:   
                employee = PersonalInfo.objects.get(fk_user=user) # handle error doesnotexist
                #employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee) 
                employee_salary = EmployeeSalary.objects.get(employee_salary_fk=employee) # handle error doesnotexist 
                attendance_list = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
                try:
                    emp_payroll_count = EmployeePayroll.objects.all().get(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff))
                except EmployeePayroll.DoesNotExist:
                    emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
                today = date.today()
                date_today = today.strftime("%m/%d/%Y") 
                context = {
                    'user': user, 
                    'cutoff': cutoff, 
                    'attendance_list':attendance_list,
                    'emp_payroll_count': emp_payroll_count,
                    'employee_salary': employee_salary,
                    'date_today': date_today,
                    'employee': employee,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }

                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))
            except EmployeeSalary.DoesNotExist: 
                return HttpResponseRedirect(reverse_lazy('application:employee_side_error_payroll_page')) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
 
def side_employee_print_payroll_page(request, id):    
    template_name = "employee_side/employee_side_print_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:  
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    
        employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee) 

        attendance_list = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
        try:
            emp_payroll_count = EmployeePayroll.objects.all().get(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff))
        except EmployeePayroll.DoesNotExist:
            emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
        today = date.today()
        date_today = today.strftime("%m/%d/%Y") 

        if request.method == 'GET':
            pass

        context = {
            'user': user, 
            'cutoff': cutoff, 
            'attendance_list':attendance_list,
            'emp_payroll_count': emp_payroll_count,
            'employee_salary': employee_salary,
            'date_today': date_today,
            'employee': employee,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 

@login_required
def side_employee_manage_leaves_page(request):
    template_name = "employee_side/employee_side_leaves_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user:
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET':
            try: 
                #employee = get_object_or_404(PersonalInfo, fk_user=user)
                employee = PersonalInfo.objects.get(fk_user=user)
                leaves = EmployeeLeaves.objects.all().filter(employee_leave_fk=employee).order_by('-id').distinct()
                context = {
                    'user': user,  
                    'employee': employee,
                    'leaves': leaves,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }

                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))  

@login_required
def side_employee_create_leave_form(request):
    template_name = "employee_side/employee_side_create_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        try:
            employee_company_details = CompanyInfo.objects.get(fk_company_user=user) 
            if request.method == 'GET':
                form = EmployeeLeavesForm(request.GET or None)
            elif request.method == 'POST':
                form = EmployeeLeavesForm(request.POST or None, request.FILES)
                if form.is_valid():
                    instance = form.save(commit=False) 
                    instance.employee_leave_fk = employee
                    instance.department = user.company_to_user.department
                    if 'attachments' in request.FILES:
                        instance.attachments = request.FILES['attachments']
                    instance.save()
                    #lookup from parent table
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
                    admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                    for admin in admins:
                        Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="You have been received leave form from {sender}".format(sender=user),category=category_list[9],level=level_list[1])
                    
                    return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_leaves_page'))
                else:
                    print("Form Error!")
            context = {
                'user': user,  
                'employee': employee, 
                'form': form,
                'employee_company_details':employee_company_details,
            }
            return render(request, template_name, context)
        except CompanyInfo.DoesNotExist:
            return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

@login_required
def side_employee_edit_leave_form(request, id):
    template_name = "employee_side/employee_side_edit_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        leave = get_object_or_404(EmployeeLeaves, id=id)
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()

        try:
            employee_company_details = CompanyInfo.objects.get(fk_company_user=user)
            if request.method == 'GET':
                form = EmployeeLeavesForm(request.GET or None, instance=leave) 
                if leave.status == 'Approved':
                    return HttpResponseRedirect(reverse_lazy('application:employee_side_leave_error_page')) 
            elif request.method == 'POST':
                form = EmployeeLeavesForm(request.POST or None, request.FILES, instance=leave)
                if form.is_valid():
                    
                    instance = form.save(commit=False)  
                    if 'attachments' in request.FILES:
                        instance.attachments = request.FILES['attachments']
                    
                    # deleting the file and the filefield
                    is_to_delete = request.POST.get('is_to_delete', "off")
                    if is_to_delete == 'on': 
                        # leave.attachments.url to use as a template to download the file
                        absolute_path = leave.attachments.path
                        #file existense verification 
                        if os.path.exists(absolute_path):
                            # delete also the file and the record on the file field
                            instance.attachments.delete(save=True)
                            if os.path.isfile(absolute_path):
                                # delete the file from the directory 
                                os.remove(absolute_path)                            
                        else:
                            print('File Not Found!')

                    instance.save()
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
                    admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                    for admin in admins:
                        Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Leave form from {sender} has been updated".format(sender=user),category=category_list[10],level=level_list[1])
                    
                    return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_leaves_page'))
                else:
                    print("Form Error!")
            context = {
                'user': user,  
                'employee': employee, 
                'leave':leave,
                'form': form,
                'notifications': notifications,
                'employee_company_details':employee_company_details,
                'notifications_count': notifications_count,
            }

            return render(request, template_name, context)
        except CompanyInfo.DoesNotExist:
            return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 

 
# Use to redirect the user if the leave form was approved
@login_required
def employee_side_leave_error_page(request):
    template_name = "employee_side/employee_side_leave_error_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET':  
            context = {
                'user':user, 
                'notifications': notifications,
                'notifications_count': notifications_count, 
            }
            return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
  
 
@login_required
def side_employee_delete_leave_form(request, id):
    data = dict()
    template_name = "employee_side/employee_side_delete_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    leave = get_object_or_404(EmployeeLeaves, id=id)

    if request.is_ajax():
        if user.is_active and user.is_staff and not user.is_superuser:
            if request.method == 'GET':
                context = {
                'user': user,  
                'employee': employee,   
                'leave': leave,          
                } 
                data['html_form'] = render_to_string(template_name, context, request)
            elif request.method == 'POST':
                leave.delete()
                data['form_is_valid'] = True 
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Leave form from {sender} has been deleted!".format(sender=user),category=category_list[11],level=level_list[1])
                
            return JsonResponse(data)
        else:
            raise Http404()
    else:
        raise Http404()

@login_required
def side_employee_view_leave_form(request, id):
    template_name="employee_side/employee_side_view_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        leave = get_object_or_404(EmployeeLeaves, id=id)
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass

        context = {
            'user': user,  
            'employee': employee,  
            'leave': leave,
            'notifications': notifications,
            'notifications_count': notifications_count,

        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
 
@login_required
def side_employee_manage_inteneraries(request):
    template_name = "employee_side/employee_side_manage_iteneraries.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        if request.method == 'GET':
            try: 
                #employee = get_object_or_404(PersonalInfo, fk_user=user)
                employee = PersonalInfo.objects.get(fk_user=user) 
                itenerary_list = EmployeeItenerary.objects.all().filter(Q(employee_itenerary_fk=employee)).order_by('-id').distinct()
                notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
                notifications_count = notifications.count()

                context = {
                    'user': user,  
                    'employee': employee, 
                    'itenerary_list': itenerary_list,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }

                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
        
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))  

@login_required
def side_employee_create_inteneraries(request):
    template_name = "employee_side/employee_side_create_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()
        EmployeeIteneraryDetailsFormset = inlineformset_factory(EmployeeItenerary, EmployeeIteneraryDetails, form=EmployeeIteneraryDetailsForm, extra=1, can_delete=False) 
            
        if request.method == 'GET':
            formEID = EmployeeIteneraryDetailsFormset(request.GET or None)
            formIF = EmployeeIteneraryForm(request.GET or None)
           
        
        elif request.method == 'POST':
            formEID = EmployeeIteneraryDetailsFormset(request.POST or None)
            formIF = EmployeeIteneraryForm(request.POST or None)

            if formEID.is_valid() and formIF.is_valid(): 
                
                instanceFIF = formIF.save(commit=False)
                instanceFIF.employee_itenerary_fk = employee
                instanceFIF.save()

                instanceFEID = formEID.save(commit=False)

                for form in instanceFEID:
                    form.employee_itenerary = instanceFIF
                    form.save()

                #lookup from parent table
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="You have been received itinerary form from {sender}".format(sender=user),category=category_list[9],level=level_list[1])
                

                return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_iteneraries_page')) 
        context = { 
            'user': user,  
            'formIF': formIF,
            'employee': employee,  
            'formEID': formEID,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
 

@login_required
def side_employee_edit_inteneraries(request, id):
    template_name = "employee_side/employee_side_edit_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:   
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        itenerary = get_object_or_404(EmployeeItenerary, id=id)
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()
        EmployeeIteneraryDetailsFormset = inlineformset_factory(EmployeeItenerary, EmployeeIteneraryDetails, form=EmployeeIteneraryDetailsForm, extra=0, can_delete=True) 
            
        if request.method == 'GET':
            formEID = EmployeeIteneraryDetailsFormset(request.GET or None, instance=itenerary) 
        
        elif request.method == 'POST':
            formEID = EmployeeIteneraryDetailsFormset(request.POST or None, instance=itenerary) 

            if formEID.is_valid():
                formEID.save()# you do not need to to loop through when deleting
                # instanceFEID = formEID.save(commit=False) 
                # for form in instanceFEID: 
                #     form.save() 
                #lookup from parent table
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Itinerary form from {sender} has been updated!".format(sender=user),category=category_list[10],level=level_list[1])
            
                return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_iteneraries_page')) 
        context = {
            'user': user,  
            'employee': employee, 
            'formEID': formEID,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 

@login_required
def side_employee_delete_inteneraries(request, id):
    data = dict()
    template_name = "employee_side/employee_side_delete_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    itenerary = get_object_or_404(EmployeeItenerary, id=id)

    if request.is_ajax():
        if user.is_active and user.is_staff and not user.is_superuser:
            if request.method == 'GET': 
                context = {
                'user': user,  
                'employee': employee,   
                'itenerary': itenerary,          
                } 
                data['html_form'] = render_to_string(template_name, context, request)
            elif request.method == 'POST':
                itenerary.delete()
                data['form_is_valid'] = True 
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Itinerary form from {sender} has been deleted!".format(sender=user),category=category_list[11],level=level_list[1])
               
            return JsonResponse(data)
        else:
            raise Http404()
    else:
        raise Http404()

@login_required
def side_employee_view_itenerary_form(request, id):
    template_name = "employee_side/employee_side_view_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:  
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        itenerary = get_object_or_404(EmployeeItenerary, id=id)
        itenerary_details = EmployeeIteneraryDetails.objects.all().filter(Q(employee_itenerary=itenerary)).order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()
    
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass 
        context = {
            'user': user,  
            'employee': employee,  
            'itenerary': itenerary,
            'itenerary_details': itenerary_details,
            'notifications': notifications,
            'notifications_count': notifications_count,
        } 
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))  

@login_required
def side_employee_manage_concerns(request):
    template_name = "employee_side/employee_side_manage_concerns.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user:
        if request.method == 'GET':
            try: 
                #employee = get_object_or_404(PersonalInfo, fk_user=user)
                employee = PersonalInfo.objects.get(fk_user=user)   
                concern_list = Concerns.objects.all().filter(Q(sender=employee)).order_by('-id').distinct()
                notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
                notifications_count = notifications.count() 

                context = {
                    'user': user,  
                    'employee': employee, 
                    'concern_list': concern_list,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }

                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))           
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))  

@login_required
def side_employee_create_concern(request):
    data = dict()
    template_name = "employee_side/employee_side_create_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)   
    if user:  
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None)   
            if form.is_valid():
                instance = form.save(commit=False)         
                instance.sender  = employee
                instance.save()

                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')
                edit_url = reverse_lazy('application:employee_side_edit_concerns_page', kwargs={'id':instance.id})
                delete_url = reverse_lazy('application:employee_side_delete_concerns_page', kwargs={'id':instance.id})
                view_url = reverse_lazy('application:employee_side_view_concerns_page', kwargs={'id':instance.id})
                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True
                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="You had received new concern from {sender}".format(sender=employee),category=category_list[9],level=level_list[1])               

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))  
    
@login_required
def side_employee_edit_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_edit_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id)
        
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)   
                instance.save()

                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')

                edit_url = reverse_lazy('application:employee_side_edit_concerns_page', kwargs={'id':concern.id})
                delete_url = reverse_lazy('application:employee_side_delete_concerns_page', kwargs={'id':concern.id})
                view_url = reverse_lazy('application:employee_side_view_concerns_page', kwargs={'id':concern.id})

                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True

                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="Concern from {sender} was updated".format(sender=employee),category=category_list[10],level=level_list[1])               


        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data) 
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 
    
@login_required
def side_employee_delete_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_delete_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id)
        if request.method == 'GET':
            context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
            } 
            data['html_form'] = render_to_string(template_name, context, request)
        elif request.method == 'POST': 
            if concern.receiver.fk_user.is_superuser:
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
            else:
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page')) 
            Notifications.objects.create(sender=user,recipient=concern.receiver.fk_user,url=url.url,message="Your concern from {sender}".format(sender=concern.sender),category=category_list[11],level=level_list[1])               
            concern.delete()
            data['form_is_valid'] = True 
        return JsonResponse(data)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 

@login_required
def side_employee_view_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_view_concern.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user:  
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id)
        if request.method == 'GET':
            pass
       
        context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 


@login_required
def side_employee_manage_inbox_concern(request):
    template_name = "employee_side/employee_side_manage_inbox_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()

        concern_list = Concerns.objects.all().filter(Q(receiver=employee)).order_by('-id').distinct()
        if request.method == 'GET':
            pass 
        context = {
            'user': user,  
            'employee': employee, 
            'concern_list': concern_list, 
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
       return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 


@login_required
def side_employee_reply_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_reply_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)   
    if user: 
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        concern = get_object_or_404(Concerns, id=id)
        if request.method == 'GET':
            form = ConcernsReplyEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsReplyEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)   
                instance.save()  

                data['form_is_valid'] = True

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 

@login_required
def side_employee_overtime_management(request):
    template_name = "employee_side/employee_side_overtime_management_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        if request.method == 'GET':
            try: 
                #employee = get_object_or_404(PersonalInfo, fk_user=user)
                employee = PersonalInfo.objects.get(fk_user=user)  
                overtime_list = Overtime.objects.all().filter(Q(employee_overtime=employee)).order_by('-id').distinct()
                notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
                notifications_count = notifications.count()

                context = {
                    'user': user,  
                    'employee': employee, 
                    'overtime_list': overtime_list,
                    'notifications': notifications,
                    'notifications_count': notifications_count,
                }

                return render(request, template_name, context)
            except PersonalInfo.DoesNotExist:
                return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))          
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page')) 

@login_required
def side_employee_create_overtime(request):
    template_name = "employee_side/employee_side_create_overtime.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user) 
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()
        OvertimeDetailsFormset = inlineformset_factory(Overtime, OvertimeDetails, form=OvertimeDetailsForm, extra=1, can_delete=False)
        try:
            employee_company_details = CompanyInfo.objects.get(fk_company_user=user) 
            if request.method == 'GET':
                formOvertime = OvertimeForm(request.GET or None) 
                formOvertimeDetails = OvertimeDetailsFormset(request.GET or None)
            elif request.method == 'POST': 
                
                formOvertime = OvertimeForm(request.POST or None) 
                formOvertimeDetails = OvertimeDetailsFormset(request.POST or None)

                if formOvertime.is_valid() and formOvertimeDetails.is_valid():

                    instanceformOvertime = formOvertime.save(commit=False)
                    instanceformOvertime.employee_overtime = employee
                    instanceformOvertime.department = employee_company_details.department 
                    instanceformOvertime.save()

                    instanceformOvertimeDetails = formOvertimeDetails.save(commit=False)

                    for form in instanceformOvertimeDetails:
                        form.overtime = instanceformOvertime
                        form.save()
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_overtime_page'))
                    admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                    for admin in admins:
                        Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="You have been received overtime form from {sender}".format(sender=user),category=category_list[9],level=level_list[1])
                    
                    return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_overtime_page'))
                else:
                    print("Error")

            context = {
                'user': user,  
                'employee': employee,  
                'notifications': notifications,
                'notifications_count': notifications_count,
                'formOvertime':formOvertime,
                'formOvertimeDetails': formOvertimeDetails,
            }

            return render(request, template_name, context)
        except CompanyInfo.DoesNotExist:
            return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

@login_required
def side_employee_edit_overtime(request, id):
    template_name = "employee_side/employee_side_edit_overtime.html" 
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)   
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        company = get_object_or_404(CompanyInfo, fk_company_user=user)
        overtime = get_object_or_404(Overtime, id=id)
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()

        OvertimeDetailsFormset = inlineformset_factory(Overtime, OvertimeDetails, form=OvertimeDetailsForm, extra=0, can_delete=True)

        if request.method == 'GET': 
            formOvertimeDetails = OvertimeDetailsFormset(request.GET or None, instance=overtime)
        elif request.method == 'POST':
            formOvertimeDetails = OvertimeDetailsFormset(request.POST or None, instance=overtime)
            if formOvertimeDetails.is_valid():
                formOvertimeDetails.save()

                url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_overtime_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Overtime form from {sender} has been updated!".format(sender=user),category=category_list[10],level=level_list[1])
               
                return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_overtime_page'))
        context = {
        'user': user,  
        'employee': employee,  
        'notifications': notifications,
        'notifications_count': notifications_count,
        #'formOvertime':formOvertime,
        'formOvertimeDetails': formOvertimeDetails,
        }

        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

@login_required
def side_employee_delete_overtime(request, id):
    data = dict()
    template_name = "employee_side/employee_side_delete_overtime.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    overtime = get_object_or_404(Overtime, id=id)

    if request.is_ajax():
        if user.is_active and user.is_staff and not user.is_superuser:
            if request.method == 'GET': 
                context = {
                'user': user,  
                'employee': employee,   
                'overtime': overtime,          
                } 
                data['html_form'] = render_to_string(template_name, context, request)
            elif request.method == 'POST':
                overtime.delete()
                data['form_is_valid'] = True 
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_overtime_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Overtime form from {sender} has been deleted!".format(sender=user),category=category_list[11],level=level_list[1])
               
            return JsonResponse(data)
        else:
            raise Http404()
    else:
        raise Http404()

@login_required
def side_employee_view_overtime_form(request, id):
    template_name = "employee_side/employee_side_view_overtime.html"
    user = get_object_or_404(User, username=request.user.username)
    user = validate(user, is_business_unit_head=True, is_employee=True)   
    if user:
        employee = get_object_or_404(PersonalInfo, fk_user=user)
        overtime = get_object_or_404(Overtime, id=id)
        overtime_details = OvertimeDetails.objects.all().filter(Q(overtime=overtime)).order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
        notifications_count = notifications.count()

        overtime_list = Overtime.objects.all().filter(id=id).order_by('-id').distinct()
        overtime_sum = OvertimeDetails.objects.all().filter(overtime__in=overtime_list).aggregate(Sum('duration')).get('duration__sum') 
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass 
        context = {
            'user': user,  
            'employee': employee,  
            'overtime': overtime,
            'overtime_details': overtime_details,
            'notifications': notifications,
            'notifications_count': notifications_count,
            'overtime_sum':overtime_sum,
        } 
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

@login_required
def employee_side_maintainance_page(request):
    template_name = "employee_side/employee_side_maintainance_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True)  
   
    if user: 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET':
            search_term = request.GET.get('search_term')
            if search_term.strip(): 
                print(search_term)

        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count,
            'search_term': search_term,
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

@login_required
def employee_side_error_page(request):
    template_name = "employee_side/employee_side_error_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True)  
    if user: 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET': 
            pass
        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count, 
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

@login_required
def employee_side_error_payroll_page(request):
    template_name = "employee_side/employee_side_error_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True) 
        
    if user: 
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
        if request.method == 'GET': 
            pass
        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count, 
        }
        return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))

@login_required
def employee_side_form_manager(request):
    template_name = "employee_side/employee_side_form_manager.html"
    user = get_object_or_404(User, username=request.user.username) 
    user = validate(user, is_business_unit_head=True, is_employee=True) 
    if user:
        if request.method == 'GET':
            notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
            notifications_count = notifications.count()

            context = {
                'user': user,   
                'notifications': notifications,
                'notifications_count': notifications_count,
            }

            return render(request, template_name, context)
    else:
        return HttpResponseRedirect(reverse_lazy('application:employee_side_error_page'))
